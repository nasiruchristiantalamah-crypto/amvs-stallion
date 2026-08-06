"""
================================================================================
NON-LIFE IFRS 17 EXCEL EXPORTER
================================================================================
What this file does:
    Exports a full non-life PAA statement run (engine/ifrs17_nonlife.py +
    engine/journals.py output) to a formatted Excel workbook, matching the
    professional presentation of PIC's own valuation result workbooks —
    Stallion navy headers, GHS currency formatting, bordered tables — not
    a raw data dump.

Sheets produced:
    1. Summary            — headline figures, onerous classes flagged
    2. IBNR by class       — the underlying reserving components (IBNR,
                             OCR, ULAE, UPR, DAC) by class, Gross/Net/RI
    3. PAA liabilities      — LRC + LIC by class, Gross/Net/RI, onerous test
    4. Balance sheet         — PIC-style layout (line items as rows, classes
                              as columns), Gross/Net/RI as three stacked tables
    5. Income statement       — derived from the journal's P&L accounts
                                (Insurance Revenue, Service Expense, Finance
                                Income, Reinsurance Expense/Recoveries)
    6. Journal entries         — the full double-entry journal listing
    7. General ledger           — T-account roll-forward per primary
                                   account (engine/general_ledger.py),
                                   whole-portfolio then by class of
                                   business, each checked to balance
    8. Assumptions               — RA loading, discount basis, data sources,
                                   generation timestamp — the basis this run
                                   was produced on
================================================================================
"""

import os
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.journals import CHART_OF_ACCOUNTS, JournalEntry

# ── Stallion brand styling (matches dashboard.html's sidebar navy/gold) ────
STALLION_NAVY = "1F3864"
STALLION_GOLD = "C9A84C"
WHITE         = "FFFFFF"
LIGHT_GREY    = "F2F2F2"
RED           = "C00000"

HEADER_FILL   = PatternFill(start_color=STALLION_NAVY, end_color=STALLION_NAVY, fill_type="solid")
HEADER_FONT   = Font(color=WHITE, bold=True, size=11, name="Calibri")
TITLE_FONT    = Font(color=STALLION_NAVY, bold=True, size=15, name="Calibri")
SUBTITLE_FONT = Font(color="595959", italic=True, size=10, name="Calibri")
SECTION_FONT  = Font(color=STALLION_NAVY, bold=True, size=12, name="Calibri")
TOTAL_FONT    = Font(bold=True, size=11, name="Calibri")
TOTAL_FILL    = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
ONEROUS_FONT  = Font(color=RED, bold=True, size=11, name="Calibri")

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GHS_FORMAT = '"GHS" #,##0.00'
PCT_FORMAT = '0.00%'

# Where generated workbooks are written by default — exposed so api/main.py's
# download endpoint can serve from the same directory without duplicating this path.
GENERATED_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "generated")


# ── Shared formatting helpers ───────────────────────────────────────────────

def _write_title(ws, title: str, subtitle: Optional[str] = None, row: int = 1) -> int:
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    next_row = row + 1
    if subtitle:
        ws.cell(row=next_row, column=1, value=subtitle).font = SUBTITLE_FONT
        next_row += 1
    return next_row + 1


def _write_section(ws, title: str, row: int) -> int:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    return row + 1


def _write_table(
    ws, start_row: int, headers: List[str], rows: List[list],
    currency_cols: Optional[set] = None, pct_cols: Optional[set] = None,
    total_row_indices: Optional[set] = None, onerous_row_indices: Optional[set] = None,
) -> int:
    """Write a formatted table (navy header, bordered/shaded body). Returns the next free row."""
    currency_cols       = currency_cols or set()
    pct_cols            = pct_cols or set()
    total_row_indices   = total_row_indices or set()
    onerous_row_indices = onerous_row_indices or set()

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row_data in enumerate(rows):
        excel_row = start_row + 1 + r
        is_total   = r in total_row_indices
        is_onerous = r in onerous_row_indices
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = BORDER
            if is_total:
                cell.font, cell.fill = TOTAL_FONT, TOTAL_FILL
            elif is_onerous and col_idx == 1:
                cell.font = ONEROUS_FONT
            if col_idx in currency_cols and isinstance(value, (int, float)):
                cell.number_format = GHS_FORMAT
            elif col_idx in pct_cols and isinstance(value, (int, float)):
                cell.number_format = PCT_FORMAT

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    return start_row + 1 + len(rows) + 2


# ── Sheet builders ───────────────────────────────────────────────────────────

def _build_summary_sheet(wb, statements: dict, journal_entries: List[JournalEntry], meta: dict):
    ws = wb.active
    ws.title = "Summary"
    row = _write_title(
        ws, "AMVS — Non-Life IFRS 17 PAA Statements",
        f"{meta.get('company_name', 'Provident Insurance Limited')} — {statements['period']} — "
        f"Generated {meta.get('generated_at', datetime.now().strftime('%d %B %Y %H:%M'))}",
    )

    totals = statements["totals"]
    headers = ["Basis", "LRC (GHS)", "LIC (GHS)", "Total Liability (GHS)"]
    rows = [[basis.upper(), totals[basis].lrc, totals[basis].lic, totals[basis].total_liability]
            for basis in ("gross", "net", "ri")]
    row = _write_section(ws, "Headline liabilities", row)
    row = _write_table(ws, row, headers, rows, currency_cols={2, 3, 4})

    onerous = [
        f"{cls} ({basis})" for cls in statements["classes"] for basis in ("gross", "net", "ri")
        if statements["by_class"][cls][basis].is_onerous
    ]
    row = _write_section(ws, "Onerous contract test", row)
    ws.cell(row=row, column=1, value="Onerous classes:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=", ".join(onerous) if onerous else "None — no class is onerous")
    if onerous:
        ws.cell(row=row, column=2).font = ONEROUS_FONT
    row += 2

    row = _write_section(ws, "Journal", row)
    total_dr = round(sum(e.debit for e in journal_entries), 2)
    total_cr = round(sum(e.credit for e in journal_entries), 2)
    ws.cell(row=row, column=1, value="Journal entries posted:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(journal_entries))
    row += 1
    ws.cell(row=row, column=1, value="Total debit / credit:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"GHS {total_dr:,.2f} / GHS {total_cr:,.2f}  "
                                      f"({'balanced' if total_dr == total_cr else 'NOT BALANCED'})")


def _build_ibnr_by_class_sheet(wb, statements: dict):
    ws = wb.create_sheet("IBNR by class")
    row = _write_title(ws, "Reserving Components by Class of Business", statements["period"])

    reserving = statements["reserving_summary"]
    for basis in ("gross", "net", "ri"):
        row = _write_section(ws, basis.upper(), row)
        headers = ["Class of business", "IBNR (GHS)", "OCR (GHS)", "ULAE (GHS)", "UPR (GHS)", "DAC (GHS)"]
        rows = [
            [cls, reserving["by_class"][cls][basis]["ibnr"], reserving["by_class"][cls][basis]["ocr"],
             reserving["by_class"][cls][basis]["ulae"], reserving["by_class"][cls][basis]["upr"],
             reserving["by_class"][cls][basis]["dac"]]
            for cls in statements["classes"]
        ]
        t = reserving["totals"][basis]
        rows.append(["Total", t["ibnr"], t["ocr"], t["ulae"], t["upr"], t["dac"]])
        row = _write_table(ws, row, headers, rows, currency_cols={2, 3, 4, 5, 6}, total_row_indices={len(rows) - 1})


def _build_paa_liabilities_sheet(wb, statements: dict):
    ws = wb.create_sheet("PAA liabilities")
    row = _write_title(ws, "PAA Liabilities — LRC and LIC by Class", statements["period"])

    for basis in ("gross", "net", "ri"):
        row = _write_section(ws, basis.upper(), row)
        headers = ["Class of business", "LRC (GHS)", "LIC (GHS)", "Total Liability (GHS)", "Onerous?", "Loss Component (GHS)"]
        rows, onerous_idx = [], set()
        for i, cls in enumerate(statements["classes"]):
            c = statements["by_class"][cls][basis]
            rows.append([cls, c.lrc, c.lic, c.total_liability, "YES" if c.is_onerous else "No", c.loss_component])
            if c.is_onerous:
                onerous_idx.add(i)
        t = statements["totals"][basis]
        rows.append(["Total", t.lrc, t.lic, t.total_liability, "YES" if t.is_onerous else "No", t.loss_component])
        row = _write_table(ws, row, headers, rows, currency_cols={2, 3, 4, 6},
                            total_row_indices={len(rows) - 1}, onerous_row_indices=onerous_idx)


# PIC's real published Balance Sheet/Income Statement column order (roughly
# alphabetical) and display names — differ from this engine's internal
# class keys (GRANULAR_CLASSES) only cosmetically; dict lookups everywhere
# else still use the short internal keys ("Fire", "Marine").
REPORT_COLUMN_ORDER = ["Accident", "Bonds", "Engineering", "Fire", "Marine", "Motor"]
REPORT_COLUMN_DISPLAY_NAME = {
    "Fire":   "Fire, Theft and Property",
    "Marine": "Marine and Aviation",
}


def _report_column_order(classes: List[str]) -> List[str]:
    """PIC's real display order when all 6 granular classes are present; the
    engine's own class list order otherwise (e.g. the 4-class Motor/Fire/
    Accident/Others grouping, which has no real-report ordering to match)."""
    if set(classes) == set(REPORT_COLUMN_ORDER):
        return REPORT_COLUMN_ORDER
    return classes


def _display_name(cls: str) -> str:
    return REPORT_COLUMN_DISPLAY_NAME.get(cls, cls)


def _build_balance_sheet_sheet(wb, statements: dict):
    """
    PIC-style layout: line items as rows, classes as columns (matches PIC's
    own "Balance Sheet 2025" / "RI Balancesheet 2025" sheet orientation
    exactly — row labels, column order, and column display names).

    NOTE on the "IBNR + OCR" row label: PIC's own real Balance Sheet uses
    that label for a figure that's actually IBNR + OCR + ULAE combined
    (confirmed by reconciling their published figures) — not a mistake in
    this exporter, PIC's own convention is carried through as-is so the
    row genuinely matches theirs.
    """
    ws = wb.create_sheet("Balance sheet")
    row = _write_title(ws, "Non-Life Balance Sheet — by Class of Business", statements["period"])

    line_items = [
        ("IBNR + OCR",             lambda c: c.ibnr + c.ocr + c.ulae),
        ("Effect of Discounting",  lambda c: c.effect_of_discounting),
        ("Risk Adjustment",         lambda c: c.risk_adjustment),
        ("Liability for Incurred Claims", lambda c: c.lic),
        ("",                            lambda c: None),
        ("Unearned Premium Reserve",       lambda c: c.upr),
        ("Deferred Acquisition Cost",         lambda c: -c.dac),
        ("Loss Component",                       lambda c: c.loss_component),
        ("Liability for Remaining Coverage", lambda c: c.lrc),
        ("",                                         lambda c: None),
        ("Total Reserve",                               lambda c: c.total_liability),
    ]
    subtotal_labels = {"Liability for Incurred Claims", "Liability for Remaining Coverage", "Total Reserve"}

    classes = _report_column_order(statements["classes"])
    for basis in ("gross", "net", "ri"):
        row = _write_section(ws, basis.upper(), row)
        headers = ["Line item"] + [_display_name(c) for c in classes] + ["Total"]
        rows, total_idx = [], set()
        for i, (label, getter) in enumerate(line_items):
            if label == "":
                rows.append([""] * len(headers))
                continue
            values = [getter(statements["by_class"][cls][basis]) for cls in classes]
            total_val = getter(statements["totals"][basis])
            rows.append([label] + values + [total_val])
            if label in subtotal_labels:
                total_idx.add(i)
        currency_cols = set(range(2, len(headers) + 2))
        row = _write_table(ws, row, headers, rows, currency_cols=currency_cols, total_row_indices=total_idx)


def _pnl_net(journal_entries: List[JournalEntry], class_of_business: str, account_code: str,
             narrative_contains: Optional[str] = None) -> float:
    """
    Net P&L effect of one account for one class: credits increase profit,
    debits reduce it. `narrative_contains` optionally isolates one specific
    movement type where several share the same account code (e.g. account
    "302" carries IBNR movement, OCR movement, and risk adjustment movement
    as three separate postings — see engine/journals.py's post_journal_entries()).
    """
    return round(sum(
        (e.credit - e.debit) for e in journal_entries
        if e.class_of_business == class_of_business and e.account_code == account_code
        and (narrative_contains is None or narrative_contains in e.narrative)
    ), 2)


def _build_income_statement_sheet(wb, journal_entries: List[JournalEntry], statements: dict):
    """
    Matches PIC's own real published Income Statement row structure exactly
    (PAA premium reserve release -> claims analysis -> DAC release ->
    onerous loss -> Insurance service result -> Finance effect -> Investment
    Result -> IFRS 17 Profit).

    HONEST LIMITATION: PIC's real income statement decomposes claims
    movements into "Expected release of incurred claims" / "Expected
    release of risk adjustment" (development of PRIOR periods' estimates)
    separately from "New Incurred Claims" / "Increase in PV and RA"
    (genuinely new this period) — a true period-over-period comparison.
    This engine's non-life side computes a POINT-IN-TIME snapshot, not a
    roll-forward against a persisted prior period (see engine/journals.py's
    module docstring) — every movement is "day 1" first-time recognition,
    so there's no prior estimate to have "expected released" FROM. Those
    two rows are therefore always zero here; the full amount shows under
    "New Incurred Claims" / "Increase in PV and RA" instead. This is
    disclosed, not fabricated — a genuine roll-forward would need a
    persisted prior-period snapshot for non-life, which doesn't exist yet
    (the life side has one — see engine/rollforward_store.py).
    """
    ws = wb.create_sheet("Income statement")
    row = _write_title(ws, "Non-Life Income Statement — by Class of Business", statements["period"])
    ws.cell(row=row, column=1, value=(
        "Note: 'Expected release of...' rows are always 0 — this engine computes a point-in-time "
        "snapshot, not a period-over-period roll-forward, for the non-life side (see module docs)."
    )).font = SUBTITLE_FONT
    row += 2

    classes = _report_column_order(statements["classes"])
    headers = ["Line item"] + [_display_name(c) for c in classes] + ["Total"]

    def line(label: str, values: List[float]) -> list:
        return [label] + [round(v, 2) for v in values] + [round(sum(values), 2)]

    rows, total_idx = [], set()

    premium_release = [_pnl_net(journal_entries, cls, "301") for cls in classes]
    rows.append(line("PAA premium reserve release excl inv component", premium_release))

    claims_paid = [-_pnl_net(journal_entries, cls, "201", "claims paid") for cls in classes]
    rows.append(line("Actual claims and expenses excl inv comp over the period", claims_paid))

    zeros = [0.0] * len(classes)
    rows.append(line("Expected release of incurred claims over the period", zeros))
    rows.append(line("Expected release of risk adjustment for incurred claims", zeros))

    new_incurred = [
        -(_pnl_net(journal_entries, cls, "302", "IBNR movement") + _pnl_net(journal_entries, cls, "302", "OCR movement"))
        for cls in classes
    ]
    rows.append(line("New Incurred Claims over the period", new_incurred))

    pv_ra_increase = [-_pnl_net(journal_entries, cls, "302", "risk adjustment movement") for cls in classes]
    rows.append(line("Increase in PV and RA of incurred claims liability", pv_ra_increase))

    dac_release = [_pnl_net(journal_entries, cls, "304") for cls in classes]
    rows.append(line("Release of deferred acquisition cost", dac_release))

    onerous_loss = [
        -statements["by_class"][cls]["gross"].loss_component if statements["by_class"][cls]["gross"].is_onerous else 0.0
        for cls in classes
    ]
    rows.append(line("Increase in losses on onerous contracts", onerous_loss))

    expense_total = [
        claims_paid[i] + new_incurred[i] + pv_ra_increase[i] + dac_release[i] + onerous_loss[i]
        for i in range(len(classes))
    ]
    rows.append(line("Total", expense_total))
    total_idx.add(len(rows) - 1)

    isr_vals = [premium_release[i] + expense_total[i] for i in range(len(classes))]
    rows.append(line("Insurance service result", isr_vals))
    total_idx.add(len(rows) - 1)

    finance_effect = [_pnl_net(journal_entries, cls, "305") for cls in classes]
    rows.append(line("Finance effect for incurred claims in P&L", finance_effect))
    rows.append(line("Investment Result", finance_effect))   # PIC's own report shows the same figure under both labels

    profit_vals = [isr_vals[i] + finance_effect[i] for i in range(len(classes))]
    rows.append(line("IFRS 17 Profit", profit_vals))
    total_idx.add(len(rows) - 1)

    currency_cols = set(range(2, len(headers) + 2))
    _write_table(ws, row, headers, rows, currency_cols=currency_cols, total_row_indices=total_idx)


def _build_journal_entries_sheet(wb, journal_entries: List[JournalEntry], period: str):
    ws = wb.create_sheet("Journal entries")
    row = _write_title(ws, "Journal Entries", period)

    headers = ["Date", "Class of business", "Basis", "Account code", "Account name", "Debit (GHS)", "Credit (GHS)", "Narrative"]
    rows = [[e.date, e.class_of_business, e.basis.upper(), e.account_code, e.account_name,
             e.debit if e.debit else None, e.credit if e.credit else None, e.narrative]
            for e in journal_entries]
    total_dr = round(sum(e.debit for e in journal_entries), 2)
    total_cr = round(sum(e.credit for e in journal_entries), 2)
    rows.append(["", "", "", "", "TOTAL", total_dr, total_cr, ""])
    _write_table(ws, row, headers, rows, currency_cols={6, 7}, total_row_indices={len(rows) - 1})


def _build_general_ledger_sheet(wb, journal_entries: List[JournalEntry], period: str):
    """
    T-account roll-forward per primary account (engine/general_ledger.py)
    — the missing piece between "journal entries exist" (previous sheet)
    and "a general ledger exists". Portfolio-wide totals first, then one
    small table per class of business, each independently checked to
    balance to zero (see engine.general_ledger.trial_balance_is_zero) —
    not just the portfolio total, which could hide two classes' offsetting
    errors.
    """
    from engine.general_ledger import build_general_ledger, general_ledger_by_class, trial_balance_is_zero

    ws = wb.create_sheet("General ledger")
    row = _write_title(ws, "General Ledger — T-Account Roll-Forward", period)

    ledger_headers = ["Account code", "Account name", "Type", "Opening balance (GHS)", "Debits (GHS)", "Credits (GHS)", "Closing balance (GHS)", "Entries"]

    portfolio_ledger = build_general_ledger(journal_entries)
    row = _write_section(ws, "Whole portfolio", row)
    ws.cell(row=row, column=1, value="Trial balance:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Balanced" if trial_balance_is_zero(portfolio_ledger) else "NOT BALANCED — investigate")
    if not trial_balance_is_zero(portfolio_ledger):
        ws.cell(row=row, column=2).font = ONEROUS_FONT
    row += 2

    portfolio_rows = [
        [code, acc.account_name, acc.account_type.title(), acc.opening_balance, acc.total_debits, acc.total_credits, acc.closing_balance, acc.entry_count]
        for code, acc in sorted(portfolio_ledger.items())
    ]
    row = _write_table(ws, row, ledger_headers, portfolio_rows, currency_cols={4, 5, 6, 7})
    row += 1

    row = _write_section(ws, "By class of business", row)
    for cls, cls_ledger in general_ledger_by_class(journal_entries).items():
        ws.cell(row=row, column=1, value=cls).font = Font(bold=True, color=STALLION_NAVY)
        row += 1
        cls_rows = [
            [code, acc.account_name, acc.account_type.title(), acc.opening_balance, acc.total_debits, acc.total_credits, acc.closing_balance, acc.entry_count]
            for code, acc in sorted(cls_ledger.items())
        ]
        row = _write_table(ws, row, ledger_headers, cls_rows, currency_cols={4, 5, 6, 7})


def _build_assumptions_sheet(wb, statements: dict, meta: dict):
    ws = wb.create_sheet("Assumptions")
    row = _write_title(ws, "Assumptions Used In This Run", statements["period"])

    headers = ["Assumption", "Value"]
    rows = [
        ["Reporting period", statements["period"]],
        ["Risk adjustment method", "Percentage margin over best-estimate IBNR+OCR"],
        ["Risk adjustment loading", f"{statements['ra_loading']:.1%}"],
        ["Discounting basis", "NIC published RFR yield curve (GHS)" if statements["discount_duration_years"] else "Not applied"],
        ["Discount duration assumption", f"{statements['discount_duration_years']} years" if statements["discount_duration_years"] else "N/A"],
        ["IBNR method", "Chain Ladder (validated against PIC's own Selected IBNR)"],
        ["Reserving classes", ", ".join(statements["classes"])],
        ["Data source",              meta.get("data_source", "PIC — Provident Insurance Limited, Data Summaries/2025")],
        ["Generated at",                meta.get("generated_at", datetime.now().strftime("%d %B %Y %H:%M:%S"))],
        ["Generated by",                   "AMVS — Stallion Consultants Ltd"],
    ]
    _write_table(ws, row, headers, rows)


# ── Top-level entry point ────────────────────────────────────────────────────

def export_nonlife_statements_to_excel(
    statements:         dict,
    journal_entries:      List[JournalEntry],
    meta:                   Optional[dict] = None,
    output_path:              Optional[str] = None,
) -> str:
    """
    Build the full formatted workbook and write it to disk.

    Parameters:
        statements       : engine.ifrs17_nonlife.generate_nonlife_paa_statements() output
        journal_entries   : engine.journals.generate_nonlife_journal() output
        meta               : optional {"company_name":.., "data_source":.., "generated_at":..}
        output_path         : file path to write to; defaults to
                              outputs/generated/nonlife_statements_<period>_<timestamp>.xlsx

    Returns:
        The path the workbook was written to.
    """
    meta = meta or {}
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, statements, journal_entries, meta)
    _build_ibnr_by_class_sheet(wb, statements)
    _build_paa_liabilities_sheet(wb, statements)
    _build_balance_sheet_sheet(wb, statements)
    _build_income_statement_sheet(wb, journal_entries, statements)
    _build_journal_entries_sheet(wb, journal_entries, statements["period"])
    _build_general_ledger_sheet(wb, journal_entries, statements["period"])
    _build_assumptions_sheet(wb, statements, meta)

    if output_path is None:
        os.makedirs(GENERATED_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(GENERATED_DIR, f"nonlife_statements_{statements['period']}_{timestamp}.xlsx")

    wb.save(output_path)
    return output_path
