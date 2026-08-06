"""
================================================================================
CUSTOM PRODUCT PRICING — Excel export
================================================================================
What this file does:
    Writes the "Download Excel" button's workbook for the dashboard's Part 4
    product pricing platform (engine/custom_pricing.py's run_custom_pricing()
    output). Sheets: Product Summary, Assumptions, Annual Cash Flow, Cash
    Flow by Covered Life (regulator-facing per-life detail — Main Life plus
    every dependant, each with their own opening lives/deaths/lapses/claims
    by rider), Reserve Projection, Profit Signature, and — when the caller
    already generated them this session — Rate Table, Sensitivity Analysis,
    and Investment Analysis. Reserve Projection, Profit Signature,
    Sensitivity Analysis, and Investment Analysis each carry a native
    embedded Excel chart alongside their data table (openpyxl.chart) —
    not just numbers, the same visualisations the dashboard itself shows.
================================================================================
"""

import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

GENERATED_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs", "generated")

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT  = Font(bold=True, size=14, color="1F3864")


def _write_title(ws, title: str, row: int = 1) -> int:
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    return row + 2


def _write_table(ws, headers: List[str], rows: List[List[Any]], start_row: int) -> int:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for r, row_vals in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=c, value=val)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
    return start_row + len(rows) + 2


def _build_summary_sheet(wb, result: dict, product_spec: dict):
    ws = wb.active
    ws.title = "Product Summary"
    row = _write_title(ws, f"Product Pricing Summary — {result.get('product_name', 'Custom Product')}")
    rows = [
        ["Product type", result.get("product_type")],
        ["Monthly premium (GHS)", result.get("monthly_premium")],
        ["Annual premium (GHS)", result.get("annual_premium")],
        ["Single premium equivalent (GHS)", result.get("single_premium_equiv")],
        ["Profit margin achieved", result.get("profit_margin")],
        ["PV Premiums (GHS)", result.get("pv_premiums")],
        ["PV Benefits (GHS)", result.get("pv_benefits")],
        ["PV Expenses (GHS)", result.get("pv_expenses")],
        ["PVFCF (GHS)", result.get("pvfcf")],
        ["Risk Adjustment (GHS)", result.get("risk_adjustment")],
        ["CSM at inception (GHS)", result.get("csm_at_inception")],
        ["LRC total (GHS)", result.get("lrc_total")],
        ["Onerous contract?", "Yes" if result.get("is_onerous") else "No"],
        ["Loss component (GHS)", result.get("loss_component")],
        ["Generated", datetime.now().strftime("%d %B %Y %H:%M")],
    ]
    _write_table(ws, ["Metric", "Value"], rows, row)


def _build_assumptions_sheet(wb, result: dict):
    ws = wb.create_sheet("Assumptions")
    a = result.get("assumptions_used", {})
    row = _write_title(ws, "Pricing Basis Used")
    rows = [["Basis name", a.get("name", "")]]
    rows.append(["Mortality loading", a.get("mortality_loading")])
    rows.append(["Gender basis", a.get("gender_main_str")])
    rows.append(["Valuation rate (p.a.)", a.get("valuation_rate_pa")])
    rows.append(["Investment return (p.a.)", a.get("investment_rate_pa")])
    rows.append(["Expense inflation (p.a.)", a.get("expense_inflation_pa")])
    rows.append(["Collection rate", a.get("collection_rate")])
    rows.append(["Target profit margin", a.get("target_profit_margin")])
    rows.append(["Policy fee (monthly)", a.get("policy_fee_monthly")])
    rows.append(["Acquisition cost", a.get("acquisition_cost")])
    rows.append(["Renewal expense (annual)", a.get("renewal_expense_annual")])
    rows.append(["Claims admin cost", a.get("claims_admin_cost")])
    rows.append(["RA method", a.get("ra_method")])
    rows.append(["CoC rate", a.get("coc_rate")])
    rows.append(["Solvency capital", a.get("solvency_capital")])
    row = _write_table(ws, ["Assumption", "Value"], rows, row)

    lapse = (a.get("lapse_schedule") or {}).get("rates", {})
    row = _write_title(ws, "Lapse schedule (annual rate by policy year)", row)
    _write_table(ws, ["Policy year", "Annual rate"], [[y, r] for y, r in lapse.items()], row)


def _build_annual_cashflow_sheet(wb, result: dict):
    ws = wb.create_sheet("Annual Cash Flow")
    row = _write_title(ws, "Annual Cash Flow Projection")
    rows = result.get("annual_cashflow", [])
    _write_table(
        ws, ["Policy year", "In-force lives", "Premium income", "Total claims", "Expenses", "Commission", "Net cash flow", "Investment income", "Cumulative profit"],
        [[r["policy_year"], r["opening_lives"], r["premium_income"], r["total_claims"], r["expenses"], r["commission"], r["net_cash_flow"], r["investment_income"], r["cumulative_profit"]] for r in rows],
        row,
    )


def _build_cashflow_by_life_sheet(wb, result: dict):
    """
    Regulator-facing detail: the SAME annual cash flows as the "Annual
    Cash Flow" sheet, but attributed to WHICH covered life they actually
    belong to (Main Life, then each dependant in turn) — not just a
    portfolio total. One section per life, all in this one sheet, each
    showing that life's own opening lives, expected deaths/lapses, and
    claims by rider. Every life's total_claims for a given year sums back
    to the combined "Annual Cash Flow" sheet's figure exactly — this is
    the same computed data attributed, not a separate calculation.
    """
    by_life = result.get("annual_cashflow_by_life") or {}
    if not by_life:
        return
    ws = wb.create_sheet("Cash Flow by Covered Life")
    row = _write_title(ws, f"Annual Cash Flow by Covered Life — {result.get('product_name', 'Custom Product')}")

    LIFE_SECTION_FILL = PatternFill(start_color="5b5967", end_color="5b5967", fill_type="solid")
    LIFE_SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")

    for life_label, rows in by_life.items():
        cell = ws.cell(row=row, column=1, value=life_label.upper())
        cell.font = LIFE_SECTION_FONT
        cell.fill = LIFE_SECTION_FILL
        last_col = max(6, len(rows[0]["claims_by_rider"]) + 5) if rows else 6
        for c in range(1, last_col + 1):
            ws.cell(row=row, column=c).fill = LIFE_SECTION_FILL
        row += 1

        rider_names = list(rows[0]["claims_by_rider"].keys()) if rows else []
        headers = ["Policy year", "Opening lives", "Expected deaths", "Expected lapses"] + \
                  [f"{r} (GHS)" for r in rider_names] + ["Total claims (GHS)"]
        table_rows = [
            [r["policy_year"], r["opening_lives"], r["expected_deaths"], r["expected_lapses"]] +
            [r["claims_by_rider"].get(rn, 0.0) for rn in rider_names] + [r["total_claims"]]
            for r in rows
        ]
        row = _write_table(ws, headers, table_rows, row)

    ws.column_dimensions["A"].width = 14


def _build_reserve_sheet(wb, result: dict):
    ws = wb.create_sheet("Reserve Projection")
    row = _write_title(ws, "Reserve Projection (Prospective Net Premium Reserve)")
    rows = result.get("reserve_projection", [])
    table_row = row
    _write_table(
        ws, ["Policy year", "Opening reserve", "Premium", "Claims", "Expenses", "Investment income", "Closing reserve"],
        [[r["policy_year"], r["opening_reserve"], r["premium"], r["claims"], r["expenses"], r["investment_income"], r["closing_reserve"]] for r in rows],
        row,
    )
    if rows:
        data_first, data_last = table_row + 1, table_row + len(rows)
        chart = LineChart()
        chart.title = "Closing reserve by policy year"
        chart.y_axis.title = "GHS"
        chart.x_axis.title = "Policy year"
        chart.add_data(Reference(ws, min_col=7, min_row=table_row, max_row=data_last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=data_first, max_row=data_last))
        ws.add_chart(chart, f"I{table_row}")


def _build_profit_signature_sheet(wb, result: dict):
    ws = wb.create_sheet("Profit Signature")
    sig = result.get("profit_signature", {})
    row = _write_title(ws, f"Profit Signature (Breakeven: Year {sig.get('breakeven_year', 'not reached')})")
    rows = sig.get("profit_by_year", [])
    table_row = row
    _write_table(ws, ["Policy year", "Profit", "Cumulative profit"], [[r["policy_year"], r["profit"], r["cumulative_profit"]] for r in rows], row)
    if rows:
        data_first, data_last = table_row + 1, table_row + len(rows)
        chart = BarChart()
        chart.title = "Profit emerging by policy year"
        chart.y_axis.title = "GHS"
        chart.x_axis.title = "Policy year"
        chart.add_data(Reference(ws, min_col=2, min_row=table_row, max_row=data_last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=data_first, max_row=data_last))
        ws.add_chart(chart, f"E{table_row}")


def _build_rate_table_sheet(wb, rate_table: Dict[int, dict]):
    ws = wb.create_sheet("Rate Table")
    row = _write_title(ws, "Premium Rate Table (Ages 18-70)")
    rows = []
    for age in sorted(rate_table.keys()):
        r = rate_table[age]
        if "error" in r:
            rows.append([age, None, None, r["error"]])
        else:
            rows.append([age, r["monthly_premium"], r["annual_premium"], "Yes" if r.get("is_onerous") else "No"])
    _write_table(ws, ["Age", "Monthly premium", "Annual premium", "Onerous?"], rows, row)


def _build_sensitivity_sheet(wb, sensitivity: List[dict]):
    ws = wb.create_sheet("Sensitivity Analysis")
    row = _write_title(ws, "Sensitivity Analysis")
    clean = [r for r in sensitivity if "error" not in r]
    rows = [[r["assumption_stressed"], r.get("stressed_premium"), r.get("difference"), r.get("pct_difference"),
              "Yes" if r.get("is_onerous") else "No", r.get("csm_at_inception")] for r in clean]
    table_row = row
    row = _write_table(ws, ["Assumption stressed", "Stressed premium", "Difference", "% difference", "Onerous?", "CSM at inception"], rows, row)
    if rows:
        data_first, data_last = table_row + 1, table_row + len(rows)
        chart = BarChart()
        chart.type = "bar"  # horizontal — tornado-style, matching the dashboard's own chart
        chart.title = "Premium impact by assumption stressed"
        chart.x_axis.title = "% change in premium"
        chart.add_data(Reference(ws, min_col=4, min_row=table_row, max_row=data_last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=data_first, max_row=data_last))
        chart.height, chart.width = 16, 24
        ws.add_chart(chart, f"H{table_row}")


def _build_investment_analysis_sheet(wb, investment_analysis: dict):
    """
    Savings/investment fund projection for endowment-type products —
    engine/investment_analysis.py's output. Per the real client workbook
    this mirrors (Afentoboa Plus's own InvestAnalysis sheet) and per
    explicit instruction, the cash flow detail here is MONTHLY, not
    annual — one column block per illustrative contribution level, plus
    a line chart of fund value growing over the term.
    """
    funds = investment_analysis.get("funds") or []
    if not funds:
        return
    ws = wb.create_sheet("Investment Analysis")
    row = _write_title(ws, "Investment Analysis — Savings Fund Projection")

    info_rows = [
        ["Risk premium (GHS/month)", investment_analysis.get("risk_premium")],
        ["Credited rate (p.a.)", investment_analysis.get("credited_rate_pa")],
        ["Term (months)", investment_analysis.get("term_months")],
    ]
    row = _write_table(ws, ["Basis", "Value"], info_rows, row)

    row = _write_title(ws, "Summary by illustrative contribution level", row)
    summary_rows = [[
        f["monthly_contribution"], investment_analysis.get("risk_premium"),
        f["summary"]["investment_portion_monthly"], f["summary"]["total_invested"],
        f["summary"]["total_interest_credited"], f["summary"]["closing_balance"],
    ] for f in funds]
    row = _write_table(
        ws, ["Contribution/month", "Risk premium", "Investment portion/month", "Total invested", "Interest credited", "Fund value at maturity"],
        summary_rows, row,
    )

    row = _write_title(ws, "Monthly fund projection (per illustrative contribution level)", row)
    table_row = row
    headers = ["Month"]
    for f in funds:
        label = f"GHS {f['monthly_contribution']}/mo"
        headers += [f"{label} — contribution", f"{label} — investment portion", f"{label} — interest credited", f"{label} — closing balance"]
    n_months = len(funds[0]["monthly_projection"])
    month_rows = []
    for m in range(n_months):
        row_vals = [funds[0]["monthly_projection"][m]["month"]]
        for f in funds:
            r = f["monthly_projection"][m]
            row_vals += [r["contribution"], r["investment_portion"], r["interest_credited"], r["closing_balance"]]
        month_rows.append(row_vals)
    _write_table(ws, headers, month_rows, row)

    if month_rows:
        data_first, data_last = table_row + 1, table_row + len(month_rows)
        chart = LineChart()
        chart.title = "Fund value growth by month"
        chart.y_axis.title = "GHS"
        chart.x_axis.title = "Month"
        for i in range(len(funds)):
            closing_col = 2 + i * 4 + 3  # Month(1) + i*4-col blocks, closing_balance is the 4th column in each block
            chart.add_data(Reference(ws, min_col=closing_col, min_row=table_row, max_row=data_last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=data_first, max_row=data_last))
        chart.height, chart.width = 16, 26
        ws.add_chart(chart, f"A{data_last + 3}")


def export_custom_pricing_to_excel(
    result:              dict,
    product_spec:        Optional[dict] = None,
    rate_table:          Optional[Dict[int, dict]] = None,
    sensitivity:         Optional[List[dict]] = None,
    investment_analysis: Optional[dict] = None,
    output_path:         Optional[str] = None,
) -> str:
    """
    Build the full custom-pricing workbook and write it to disk.

    Parameters:
        result              : engine.custom_pricing.run_custom_pricing() output
        product_spec        : the raw CustomProductRequest dict, for context (currently only product_name is used beyond what `result` already carries)
        rate_table          : optional run_custom_rate_table() output — adds a Rate Table sheet if given
        sensitivity         : optional run_custom_sensitivity() output — adds a Sensitivity Analysis sheet if given
        investment_analysis : optional /pricing/custom/investment-analysis response — adds an Investment Analysis sheet if given (endowment-type products only)
        output_path         : file path to write to; defaults to outputs/generated/CustomPricing_<name>_<timestamp>.xlsx

    Returns:
        The path the workbook was written to.
    """
    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, result, product_spec or {})
    _build_assumptions_sheet(wb, result)
    _build_annual_cashflow_sheet(wb, result)
    _build_cashflow_by_life_sheet(wb, result)
    _build_reserve_sheet(wb, result)
    _build_profit_signature_sheet(wb, result)
    if rate_table:
        _build_rate_table_sheet(wb, rate_table)
    if sensitivity:
        _build_sensitivity_sheet(wb, sensitivity)
    if investment_analysis:
        _build_investment_analysis_sheet(wb, investment_analysis)

    if output_path is None:
        os.makedirs(GENERATED_DIR, exist_ok=True)
        name_slug = "".join(c if c.isalnum() else "_" for c in result.get("product_name", "CustomProduct"))[:40]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(GENERATED_DIR, f"CustomPricing_{name_slug}_{timestamp}.xlsx")

    wb.save(output_path)
    return output_path
