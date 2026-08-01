"""
================================================================================
LIFE (GMM) IFRS 17 EXCEL EXPORTER
================================================================================
What this file does:
    Exports a single run_ifrs17() result to a formatted Excel workbook —
    the life-side counterpart to outputs/excel_exporter.py's non-life PAA
    exporter, reusing its Stallion navy/gold styling helpers so both look
    like one system rather than two different tools.

Sheets produced:
    1. Summary                        — company, period, product, measurement
                                          model, currency, headline liabilities
    2. Insurance Contract Liabilities  — LRC (PVFCF/RA/CSM) and LIC (best
                                          estimate/RA)
    3. Income Statement                 — insurance revenue, expenses,
                                          insurance service result
    4. CSM Roll-Forward                  — opening -> accretion -> amortisation -> closing
    5. Solvency                           — available/required capital, CAR,
                                          solvency status

Input shape:
    Takes the same flattened dict api/main.py's POST /ifrs17 endpoint
    returns as "data" (see _build_ifrs17_response_data() in api/main.py) —
    both routes are built from one shared helper so the numbers in the
    JSON response and this workbook can never drift apart.
================================================================================
"""

import os
from datetime import datetime
from typing import Optional

import openpyxl

from outputs.excel_exporter import GENERATED_DIR, _write_section, _write_table, _write_title


def export_ifrs17_to_excel(data: dict, meta: Optional[dict] = None, output_path: Optional[str] = None) -> str:
    """
    Build the formatted workbook and write it to disk.

    Parameters:
        data        : the flattened dict returned as POST /ifrs17's "data"
                      (period, company, product, measurement_model, currency,
                      lrc{}, lic{}, total_liabilities, total_liabilities_usd,
                      pnl{}, csm_rollforward{}, solvency{})
        meta        : optional {"generated_at": ...}
        output_path : file path to write to; defaults to
                      outputs/generated/ifrs17_<company>_<period>_<timestamp>.xlsx

    Returns:
        The path the workbook was written to.
    """
    meta = meta or {}
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    row = _write_title(
        ws, f"AMVS — IFRS 17 Valuation — {data['company']}",
        f"{data['product']} — {data['period']} — "
        f"Generated {meta.get('generated_at', datetime.now().strftime('%d %B %Y %H:%M'))}",
    )
    row = _write_section(ws, "Headline", row)
    _write_table(ws, row, ["Metric", "Value"], [
        ["Measurement model",        data["measurement_model"]],
        ["Currency",                 data["currency"]],
        ["In-force policies",        data["in_force_count"]],
        ["Total liabilities (GHS)",  data["total_liabilities"]],
        ["Total liabilities (USD)",  data["total_liabilities_usd"]],
    ], currency_cols={2}, total_row_indices={3})

    ws2 = wb.create_sheet("Contract Liabilities")
    row = _write_title(ws2, "Insurance Contract Liabilities", data["period"])
    row = _write_section(ws2, "LRC — Liability for Remaining Coverage (GMM)", row)
    lrc = data["lrc"]
    row = _write_table(ws2, row, ["Component", "GHS"], [
        ["PVFCF (Building Block 1)",           lrc["pvfcf"]],
        ["Risk Adjustment (Building Block 2)", lrc["risk_adjustment"]],
        ["CSM (Building Block 3)",             lrc["csm"]],
        ["Total LRC",                          lrc["total"]],
    ], currency_cols={2}, total_row_indices={3})
    row = _write_section(ws2, "LIC — Liability for Incurred Claims", row)
    lic = data["lic"]
    _write_table(ws2, row, ["Component", "GHS"], [
        ["Best Estimate",   lic["best_estimate"]],
        ["Risk Adjustment", lic["risk_adjustment"]],
        ["Total LIC",       lic["total"]],
    ], currency_cols={2}, total_row_indices={2})

    ws3 = wb.create_sheet("Income Statement")
    row = _write_title(ws3, "IFRS 17 Income Statement", data["period"])
    pnl = data["pnl"]
    _write_table(ws3, row, ["Metric", "GHS"], [
        ["Insurance Revenue",             pnl["insurance_revenue"]],
        ["Insurance Expenses",            pnl["insurance_expenses"]],
        ["Insurance Service Result",      pnl["insurance_service_result"]],
        ["  — of which CSM Amortisation", pnl["csm_amortisation"]],
        ["  — of which RA Release",       pnl["ra_release"]],
    ], currency_cols={2}, total_row_indices={2})

    ws4 = wb.create_sheet("CSM Roll-forward")
    row = _write_title(ws4, "CSM Roll-Forward (IFRS 17 Para. 44)", data["period"])
    csm = data["csm_rollforward"]
    _write_table(ws4, row, ["Movement", "GHS"], [
        ["Opening CSM",         csm["opening"]],
        ["Interest Accretion",  csm["interest_accretion"]],
        ["CSM Amortisation",    csm["amortisation"]],
        ["Closing CSM",         csm["closing"]],
    ], currency_cols={2}, total_row_indices={0, 3})

    ws5 = wb.create_sheet("Solvency")
    row = _write_title(ws5, "Solvency and Capital Adequacy (NIC GIRBC)", data["period"])
    sol = data["solvency"]
    _write_table(ws5, row, ["Metric", "Value"], [
        ["Available Capital (GHS)", sol["available_capital"]],
        ["Required Capital (GHS)",  sol["required_capital"]],
        ["Capital Adequacy Ratio",  f"{sol['capital_adequacy_ratio']:.1%}"],
        ["Solvency Status",         "SOLVENT" if sol["is_solvent"] else "CAPITAL BREACH"],
    ], currency_cols={2})

    if output_path is None:
        os.makedirs(GENERATED_DIR, exist_ok=True)
        company_slug = "".join(c if c.isalnum() else "_" for c in data["company"])[:40]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(GENERATED_DIR, f"ifrs17_{company_slug}_{data['period']}_{timestamp}.xlsx")

    wb.save(output_path)
    return output_path
