"""
================================================================================
DATA LOADER — multi-client non-life source workbooks
================================================================================
What this file does:
    Reads a client's own Excel workbooks directly and returns structured
    Python data ready to feed engine.chain_ladder / engine.bornhuetter_ferguson
    (via engine.runner.run_reserving) plus the OCR, ULAE, UPR and DAC figures
    needed for the full NIC non-life summary table (see run_nic_summary() in
    engine/runner.py).

    This is the non-life equivalent of hand-typing assumptions into
    engine/assumptions.py on the life side — except here the "assumptions"
    are the client's own live claims/premium data, so we read it directly
    instead of re-keying it.

Multi-client — IMPORTANT:
    Every function here takes a `client_id`, resolved via engine/clients.py
    to a ClientConfig. Two files feed that config: the FOLDER a client's
    workbooks live in (data_folder — either the <CLIENT_ID>_DATA_DIR
    environment variable in production, or clients/<id>/client.yaml
    locally, gitignored since it's a real machine-specific path — see
    engine/clients.py's module docstring) and the workbook/sheet NAMES
    within that folder (data_files — clients/<id>/assumptions.yaml,
    committed to git since filenames aren't sensitive).
    Originally this module was hardcoded to PIC's exact file paths; it was
    generalized after finding that a second client (QIC) uses the SAME
    workpaper template (same sheet names: MOTOR/FIRE/ACCIDENT/OTHERS,
    Premium & Commission-2025/Paid Claims-2025/Outstanding Claims-2025,
    ULAE Calculation (Rev)) but different FILE names and, in one sheet
    (Outstanding Claims), a genuinely different COLUMN layout (extra
    columns, Policy/Claim Number order swapped, amount columns shifted).
    Two things are therefore client-configurable:
        1. File and sheet names          — clients/<id>/assumptions.yaml's data_files
        2. Column positions within a sheet — NOT configurable; instead, every
           parser below locates columns by searching for their header TEXT
           rather than assuming a fixed index, so the same code handles both
           PIC's and QIC's layouts without per-client column-offset config.

Source workbooks used (see a client's assumptions.yaml data_files: section
for the exact names — these are PIC's, the default every other client's
data_files merges on top of):
    "2025 IBNR Projection (Gross & Net) - Final.xlsx"  (ibnr_workbook)
        Sheets MOTOR / FIRE / ACCIDENT / OTHERS — cumulative incurred claims
        triangle (Gross & Net) and earned/written premium by underwriting
        year, per class. This is the client's own reserving workpaper — see
        engine/chain_ladder.py and engine/bornhuetter_ferguson.py for the
        validation of this module's output against PIC's real figures.
    "2025 PIC Final Data.xlsx"  (raw_data_workbook)
        A sheet whose name CONTAINS "Outstanding Claims" (matched by
        substring, not an exact "Outstanding Claims-2025" — see
        _read_sheet_rows_containing()) — raw case reserve (OCR) register,
        one row per open claim, with Gross/Net Amount Outstanding.
    "UPR & DAC (2025).xlsx"  (upr_dac_workbook / upr_dac_sheet)
        Gross/Net Unearned Premium Reserve and Deferred Acquisition Cost by
        class of business (already computed by the client).
    "PIC ULAE (2025) - Final.xlsx"  (ulae_workbook / ulae_sheet)
        Unallocated Loss Adjustment Expense by class of business
        (Alpha-ratio method; Gross basis only — ULAE is not reinsured, so
        Net ULAE = Gross ULAE and RI ULAE = 0, per PIC's own reporting
        convention — assumed to hold for other clients too unless told
        otherwise).

Bucket structure — IMPORTANT:
    The claims triangles (and therefore IBNR) only exist for 4 classes in
    the IBNR Projection workbook: Motor, Fire, Accident, Others. The other
    source workbooks (OCR, ULAE, UPR/DAC) are more granular — they break out
    Bonds, Engineering, Liability, Marine and Aviation, etc. separately.
    This loader aggregates all non-Motor/Fire/Accident classes into a single
    "Others" bucket (matching the IBNR Projection workbook's own "OTHERS"
    sheet grouping) so every data source lines up against the same 4-class
    structure for the NIC summary table.

Structural parsing, not hardcoded row numbers:
    Row positions differ across sheets (e.g. Others' tables are shorter
    because its 2018 origin year is empty). Every table is located by its
    header text ("Underwriting Year" / "Cummulative Loss Reported" /
    "Premium" / "Class of Business" etc.), not fixed row/column numbers.
    One structural constant IS relied on in the triangle/premium tables:
    PIC and QIC's shared template consistently places the Net (reinsurance)
    block exactly 13 columns to the right of the matching Gross block — a
    ValueError is raised if that offset doesn't check out, rather than
    silently parsing the wrong numbers.
================================================================================
"""

import os
from dataclasses import replace
from typing import Dict, List, Optional, Tuple

import openpyxl

from engine.clients import ClientConfig, load_client

RESERVING_CLASSES = ["Motor", "Fire", "Accident", "Others"]

# The 6-class breakdown PIC's (and presumably most Ghanaian non-life
# insurers') own published reports actually use — "Others" above is a
# reserving-engine-only grouping forced by the IBNR triangle workbook only
# having 4 sheets (see GRANULAR_CLASSES section below for why IBNR alone
# can't be split this finely without an allocation step).
GRANULAR_CLASSES = ["Motor", "Fire", "Accident", "Bonds", "Engineering", "Marine"]

# Net (reinsurance) blocks sit this many columns to the right of the
# matching Gross block, consistently, in the shared workbook template.
NET_COLUMN_OFFSET = 13


# ── Class-of-business bucketing (4-class — drives reserving/IBNR) ──────────
#
# The raw source workbooks don't use one consistent taxonomy — the
# Outstanding Claims register alone uses product-level labels like "GROUP
# PERSONAL ACCIDENT", "ASSETS ALL RISK", "PLANT AND MACHINERY" that don't
# exactly match "Accident"/"Fire, Theft and Property" as spelled in the UPR
# & DAC / ULAE workbooks. Matching is therefore by keyword (substring),
# case-insensitive, not exact string — anything not containing "motor",
# "fire", or "accident" falls into "Others", matching the IBNR Projection
# workbook's own "OTHERS" catch-all grouping (Bonds, Engineering,
# Liability, Marine, Aviation, Workmen's Compensation, etc.).

def _bucket(raw_label: str) -> str:
    if not raw_label:
        return "Others"
    label = str(raw_label).strip().lower()
    if "motor" in label:
        return "Motor"
    if "fire" in label:
        return "Fire"
    if "accident" in label:
        return "Accident"
    return "Others"


# ── Class-of-business bucketing (6-class GRANULAR — matches published reports) ─
#
# PIC's UPR & DAC and ULAE workbooks already store Bonds/Engineering/Marine
# as their OWN rows, spelled slightly differently from the report
# ("Fire, Theft and Property" -> "Fire", "Marine and Aviation" -> "Marine")
# — _GRANULAR_LABEL_ALIASES below normalises those, confirmed by matching
# every class total dollar-for-dollar against PIC's real published Balance
# Sheet (see clients/pic/assumptions.yaml's ocr_class_mapping comment).
#
# The Outstanding Claims (OCR) register is the odd one out: it uses
# PRODUCT-level labels ("PLANT AND MACHINERY", "GOODS IN TRANSIT", "ASSETS
# ALL RISK", ...) that don't correspond to report classes by keyword at
# all — e.g. "GOODS IN TRANSIT" (which sounds like Marine cargo) actually
# reports under PIC's "Accident" class, and "PLANT AND MACHINERY" reports
# under "Engineering", not "Motor". This mapping is genuinely
# insurer-specific (each company's own product-to-class convention), so it
# is NOT guessed by keyword here — it must be configured per client (see
# ClientConfig.ocr_class_mapping, sourced from assumptions.yaml). A raw
# label with no configured mapping and no obvious Motor/Fire/Accident
# keyword match falls into "Unclassified" rather than being silently
# dropped or guessed into the wrong class.
_GRANULAR_LABEL_ALIASES = {
    "fire, theft and property": "Fire",
    "marine and aviation":      "Marine",
    "accident":                 "Accident",
    "bonds":                    "Bonds",
    "engineering":               "Engineering",
    "motor":                     "Motor",
}


def _bucket_granular(raw_label: str, ocr_class_mapping: Optional[Dict[str, str]] = None) -> str:
    """
    Resolve a raw class-of-business label to one of GRANULAR_CLASSES, or
    "Unclassified" if it can't be resolved — see module docstring above.
    Tries, in order: (1) the known UPR/DAC/ULAE label spellings, (2) a
    client-configured product->class mapping (for OCR's product-level
    labels), (3) a generic Motor/Fire/Accident keyword fallback.
    """
    if not raw_label:
        return "Unclassified"
    label = str(raw_label).strip().lower()

    if label in _GRANULAR_LABEL_ALIASES:
        return _GRANULAR_LABEL_ALIASES[label]

    if ocr_class_mapping:
        mapped = ocr_class_mapping.get(label) or ocr_class_mapping.get(str(raw_label).strip())
        if mapped:
            return mapped

    if "motor" in label:
        return "Motor"
    if "fire" in label:
        return "Fire"
    if "accident" in label:
        return "Accident"
    return "Unclassified"


# ── Workbook helpers ─────────────────────────────────────────────────────────

def _resolve_client(client_id: str, data_folder_override: Optional[str] = None) -> ClientConfig:
    """
    Resolve a client's config, optionally reading its workbooks from a
    different folder — e.g. a temporary directory of files a user just
    uploaded through the dashboard (see api/main.py's /upload/* endpoints)
    — while keeping that client's own data_files naming convention and
    ocr_class_mapping. The uploaded files must be named exactly as that
    client's assumptions.yaml data_files expects (PIC's own filenames, if
    client_id="pic" — the only validated non-life workbook template so far).
    """
    client = load_client(client_id)
    if data_folder_override:
        client = replace(client, data_folder=data_folder_override)
    return client


def _read_sheet_rows(client: ClientConfig, data_file_key: str, sheet_name: str) -> List[tuple]:
    path = client.data_file_path(data_file_key)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {path} for client '{client.client_id}' "
                f"— available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _read_sheet_rows_containing(client: ClientConfig, data_file_key: str, name_contains: str) -> List[tuple]:
    """
    Like _read_sheet_rows(), but matches the first sheet whose NAME CONTAINS
    `name_contains` (case-insensitive) rather than requiring an exact sheet
    name — for sheets whose real name carries a reporting year/period suffix
    that changes every run (e.g. "Outstanding Claims-2025" this year,
    "Outstanding Claims-2026" next year) and shouldn't need a code change
    just because the year rolled over.
    """
    path = client.data_file_path(data_file_key)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        match = next((s for s in wb.sheetnames if name_contains.lower() in s.lower()), None)
        if match is None:
            raise ValueError(
                f"No sheet containing '{name_contains}' found in {path} for client '{client.client_id}' "
                f"— available sheets: {wb.sheetnames}"
            )
        ws = wb[match]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _as_year(value) -> Optional[int]:
    """Return value as an int year if it parses cleanly, else None."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _cell(row: tuple, col: int):
    return row[col] if 0 <= col < len(row) else None


def _find_col(row: tuple, label: str, mode: str = "exact") -> Optional[int]:
    """
    Find the column index of a cell matching `label` in this row.
    mode: "exact", "startswith", or "contains" (case-sensitive — header
    text is stable enough within one workbook family that this hasn't
    needed to be case-insensitive so far).
    """
    for col, val in enumerate(row):
        if not isinstance(val, str):
            continue
        if mode == "exact" and val == label:
            return col
        if mode == "startswith" and val.startswith(label):
            return col
        if mode == "contains" and label in val:
            return col
    return None


def _find_row_with_col(rows: List[tuple], label: str, mode: str = "exact") -> Tuple[Optional[int], Optional[int]]:
    """Find the first row containing a cell matching `label`; returns (row_idx, col_idx) or (None, None)."""
    for i, row in enumerate(rows):
        col = _find_col(row, label, mode)
        if col is not None:
            return i, col
    return None, None


# ── 1. Claims triangles + earned/written premium (drives IBNR) ─────────────

def load_triangle(class_of_business: str, client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, Dict[int, float]]:
    """
    Read the cumulative Gross/Net incurred claims triangle and Gross/Net
    earned/written premium for one of the 4 reserving classes from the
    client's IBNR Projection workbook.

    Parameters:
        class_of_business    : "Motor", "Fire", "Accident", or "Others"
        client_id              : which client (clients/<client_id>/assumptions.yaml)
        data_folder_override  : read this client's workbooks from a different
                                  folder (e.g. an uploaded-files temp dir) — see
                                  _resolve_client()

    Returns:
        {
            "gross_triangle": {origin_year: [cumulative incurred by age...]},
            "net_triangle":   {origin_year: [...]},
            "gross_premium":  {origin_year: float},
            "net_premium":    {origin_year: float},
        }
        Ready to pass straight into engine.runner.run_reserving().
    """
    if class_of_business not in RESERVING_CLASSES:
        raise ValueError(f"Unknown reserving class '{class_of_business}' — must be one of {RESERVING_CLASSES}")

    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "ibnr_workbook", class_of_business.upper())

    gross_triangle, net_triangle = _parse_triangle_table(rows)
    gross_premium,  net_premium  = _parse_premium_table(rows)

    return {
        "gross_triangle": gross_triangle,
        "net_triangle":   net_triangle,
        "gross_premium":  gross_premium,
        "net_premium":    net_premium,
    }


def _parse_triangle_table(rows: List[tuple]) -> Tuple[Dict[int, List[float]], Dict[int, List[float]]]:
    """Locate and parse the "Cummulative Loss Reported" table (Gross + Net blocks)."""
    header_idx = None
    gross_col  = None
    for i, row in enumerate(rows):
        for col in range(len(row) - 1):
            label = _cell(row, col)
            marker = _cell(row, col + 1)
            if label == "Underwriting Year" and isinstance(marker, str) and marker.startswith("Cummulative Loss"):
                header_idx, gross_col = i, col
                break
        if header_idx is not None:
            break

    if header_idx is None:
        raise ValueError("Could not locate 'Cummulative Loss Reported' table — sheet layout may have changed")

    net_col = gross_col + NET_COLUMN_OFFSET
    net_label = _cell(rows[header_idx], net_col)
    if net_label not in ("Underwriting Year", "Accident Year"):
        raise ValueError(
            f"Expected the Net block's year label at column {net_col} (offset {NET_COLUMN_OFFSET} "
            f"from the Gross block at column {gross_col}) but found {net_label!r} — layout mismatch."
        )

    # Data rows start 2 rows below the header (row header+1 is the age-index sub-header).
    # The template includes one extra blank placeholder row after the last real
    # origin year (e.g. a 2026 row with no age-0 value at all) — stop there, don't
    # include it, since "year parses as int" alone isn't enough to detect it.
    origin_years: List[int] = []
    data_start = header_idx + 2
    for row in rows[data_start:]:
        year = _as_year(_cell(row, gross_col))
        if year is None or _cell(row, gross_col + 1) is None:
            break
        origin_years.append(year)

    if not origin_years:
        raise ValueError("No origin years found under the Cummulative Loss Reported table")

    max_year = max(origin_years)
    gross_triangle: Dict[int, List[float]] = {}
    net_triangle:   Dict[int, List[float]] = {}

    for offset, year in enumerate(origin_years):
        row = rows[data_start + offset]
        n_periods = max_year - year + 1  # diagonal rule: fewer periods for more recent origin years
        gross_triangle[year] = [float(_cell(row, gross_col + 1 + k) or 0.0) for k in range(n_periods)]
        net_triangle[year]   = [float(_cell(row, net_col   + 1 + k) or 0.0) for k in range(n_periods)]

    return gross_triangle, net_triangle


def _parse_premium_table(rows: List[tuple]) -> Tuple[Dict[int, float], Dict[int, float]]:
    """
    Locate and parse the first "Expected Loss Ratio" premium table (Earned
    or Written Premium — whichever label the client used for this class).
    Some class sheets carry a second, unused, all-zero ELR section further
    down the sheet — we deliberately take the FIRST occurrence, which is
    the one the client's own Selected IBNR is actually derived from.
    """
    header_idx = None
    year_col   = None
    for i, row in enumerate(rows):
        for col in range(len(row) - 3):
            year_label     = _cell(row, col)
            reported_label = _cell(row, col + 1)
            premium_label  = _cell(row, col + 2)
            if (year_label in ("Underwriting Year", "Accident Year")
                    and isinstance(reported_label, str) and reported_label.startswith("Reported")
                    and isinstance(premium_label, str) and "Premium" in premium_label):
                header_idx, year_col = i, col
                break
        if header_idx is not None:
            break

    if header_idx is None:
        raise ValueError("Could not locate an Earned/Written Premium table — sheet layout may have changed")

    premium_col = year_col + 2
    net_year_col = year_col + NET_COLUMN_OFFSET
    net_premium_col = premium_col + NET_COLUMN_OFFSET
    net_label = _cell(rows[header_idx], net_year_col)
    if net_label not in ("Underwriting Year", "Accident Year"):
        raise ValueError(
            f"Expected the Net premium block's year label at column {net_year_col} "
            f"but found {net_label!r} — layout mismatch."
        )

    # Data rows start 2 rows below the header — row header+1 is a sub-header
    # splitting "Claim Ratio" into "Actual"/"Selected" (same pattern as the
    # age-index sub-header in the triangle table above).
    gross_premium: Dict[int, float] = {}
    net_premium:   Dict[int, float] = {}
    for row in rows[header_idx + 2:]:
        year = _as_year(_cell(row, year_col))
        if year is None:
            break
        gross_premium[year] = float(_cell(row, premium_col) or 0.0)
        net_premium[year]   = float(_cell(row, net_premium_col) or 0.0)

    return gross_premium, net_premium


# ── 1b. IBNR / URR split, by class (optional — see docstring) ──────────────

def load_ibnr_urr_split(client_id: str = "pic", basis: str = "gross", data_folder_override: Optional[str] = None) -> Dict[str, Dict[str, float]]:
    """
    Read the client's own IBNR/URR split, if they produce one — see
    clients/pic/assumptions.yaml's bel_workbook comment for what URR
    (Unexpired Risk Reserve) is and why it needs excluding from LIC. The
    workbook this reads has the same Motor/Fire/Accident/Others sheet
    layout as the IBNR triangle workbook, with a "Total" row per class
    holding columns (0-indexed, values only — this workbook's layout is
    more idiosyncratic than the others, so positions are relied on
    directly rather than searched for by header text): 0=Underwriting Year
    label ("Total"), 1=GWP, 2=OS, 3=IBNR+URR combined, 4=IBNR (pure),
    5=URR, 6=Paid Claims, 7=Ult Loss Ratios.

    Parameters:
        basis : "gross" (bel_workbook) or "net" (bel_net_workbook)

    Raises ValueError if this client hasn't configured the relevant
    workbook — callers should catch this and fall back to the combined
    chain-ladder IBNR+URR figure; not every client will produce this
    refinement.

    Returns:
        {"Motor": {"ibnr": float, "urr": float}, "Fire": {...}, ...}
    """
    if basis not in ("gross", "net"):
        raise ValueError(f"basis must be 'gross' or 'net', got {basis!r}")
    data_key = "bel_workbook" if basis == "gross" else "bel_net_workbook"

    client = _resolve_client(client_id, data_folder_override)
    if data_key not in client.data_files or not os.path.isfile(client.data_file_path(data_key)):
        # Same fallback whether the key is genuinely unconfigured (a client
        # that doesn't produce this split at all) or configured-but-missing
        # on disk (e.g. an upload flow — see api/main.py's /upload/* — that
        # reuses another client's data_files template but only received the
        # 4 required workbooks, not this optional refinement).
        raise ValueError(
            f"Client '{client_id}' has no '{data_key}' available — this client doesn't "
            f"produce an IBNR/URR split; fall back to the combined chain-ladder IBNR+URR figure instead."
        )

    result: Dict[str, Dict[str, float]] = {}
    for cls in RESERVING_CLASSES:
        rows = _read_sheet_rows(client, data_key, cls)
        for row in rows:
            if _cell(row, 0) == "Total":
                result[cls] = {"ibnr": float(_cell(row, 4) or 0.0), "urr": float(_cell(row, 5) or 0.0)}
                break
        else:
            raise ValueError(f"Could not find a 'Total' row in the '{cls}' sheet of {client.data_files[data_key]}")
    return result


# ── 2. Outstanding claims / OCR (case reserves), by class ──────────────────

def load_ocr(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, Dict[str, float]]:
    """
    Aggregate the Outstanding Claims register (one row per open claim)
    into Gross/Net OCR totals per reserving bucket.

    Column positions are located by header text, not fixed indices — PIC's
    and QIC's own registers lay this sheet out differently (different
    column order/count, an extra "IFRS 17 Portfolio" column on QIC's) but
    both share the same 2-row header shape: a group-header row ("Gross
    Amount Outstanding" / "Net Amount Outstanding") sits directly ABOVE the
    field-header row ("Claim Number" / "Policy Number" / "Class of
    Business" / ...).

    Amounts are stored as negative numbers in the raw export (a
    ledger-credit sign convention) — this loader takes the absolute value
    so OCR is reported as a positive reserve, matching the other figures.

    Returns:
        {"Motor": {"gross": float, "net": float}, "Fire": {...}, ...}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows, class_col, gross_col, net_col, field_header_idx = _parse_ocr_sheet(client)

    totals = {c: {"gross": 0.0, "net": 0.0} for c in RESERVING_CLASSES}
    for row in rows[field_header_idx + 1:]:
        class_label = _cell(row, class_col)
        if not class_label:
            continue
        bucket = _bucket(class_label)
        totals[bucket]["gross"] += abs(float(_cell(row, gross_col) or 0.0))
        totals[bucket]["net"]   += abs(float(_cell(row, net_col) or 0.0))

    return totals


def _parse_ocr_sheet(client: ClientConfig) -> Tuple[List[tuple], int, int, int, int]:
    """Shared header/column detection for load_ocr() and load_ocr_granular() — returns (rows, class_col, gross_col, net_col, field_header_idx)."""
    # Matched by name CONTAINING "Outstanding Claims", not an exact
    # "Outstanding Claims-2025" — the real sheet name carries the reporting
    # year as a suffix, which changes every run (see
    # _read_sheet_rows_containing()'s docstring for why this can't be a
    # hardcoded literal without breaking every year the report rolls over).
    rows = _read_sheet_rows_containing(client, "raw_data_workbook", "Outstanding Claims")

    field_header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if field_header_idx is None:
        raise ValueError("Could not locate a 'Class of Business' header in the Outstanding Claims sheet")
    if _find_col(rows[field_header_idx], "Policy Number", mode="exact") is None:
        raise ValueError("Found 'Class of Business' but not 'Policy Number' in the same row — unexpected sheet layout")

    if field_header_idx == 0:
        raise ValueError("Expected a group-header row above the field-header row, but the header is the first row")
    group_header_row = rows[field_header_idx - 1]
    gross_col = _find_col(group_header_row, "Gross Amount Outstand", mode="startswith")
    net_col   = _find_col(group_header_row, "Net Amount Outstand", mode="startswith")
    if gross_col is None or net_col is None:
        raise ValueError(
            "Could not locate 'Gross Amount Outstanding' / 'Net Amount Outstanding' in the row "
            "above the Outstanding Claims header — unexpected sheet layout"
        )
    return rows, class_col, gross_col, net_col, field_header_idx


def load_ocr_granular(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, Dict[str, float]]:
    """
    Same source data as load_ocr(), aggregated into the 6-class breakdown
    (GRANULAR_CLASSES) PIC's real published reports use, plus an
    "Unclassified" bucket for any raw label that can't be resolved (see
    _bucket_granular()) — surfaced rather than silently dropped or
    misallocated, so an incomplete client mapping is visible in the output.

    Returns:
        {"Motor": {"gross": float, "net": float}, ..., "Unclassified": {...}}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows, class_col, gross_col, net_col, field_header_idx = _parse_ocr_sheet(client)

    totals = {c: {"gross": 0.0, "net": 0.0} for c in GRANULAR_CLASSES + ["Unclassified"]}
    for row in rows[field_header_idx + 1:]:
        class_label = _cell(row, class_col)
        if not class_label:
            continue
        bucket = _bucket_granular(class_label, client.ocr_class_mapping)
        totals[bucket]["gross"] += abs(float(_cell(row, gross_col) or 0.0))
        totals[bucket]["net"]   += abs(float(_cell(row, net_col) or 0.0))

    return totals


# ── 3. UPR & DAC, by class ──────────────────────────────────────────────────

def load_upr_dac(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, Dict[str, float]]:
    """
    Read the client's own computed Gross/Net UPR and DAC by class, and
    aggregate into the 4 reserving buckets.

    Returns:
        {"Motor": {"gross_upr":.., "net_upr":.., "gross_dac":.., "net_dac":..}, ...}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "upr_dac_workbook", client.data_files["upr_dac_sheet"])

    header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if header_idx is None:
        raise ValueError("Could not locate the 'Class of Business' header row in the UPR & DAC sheet")
    gross_upr_col = class_col + 1
    net_upr_col   = class_col + 2
    gross_dac_col = class_col + 3
    net_dac_col   = class_col + 4

    totals = {c: {"gross_upr": 0.0, "net_upr": 0.0, "gross_dac": 0.0, "net_dac": 0.0} for c in RESERVING_CLASSES}
    for row in rows[header_idx + 1:]:
        label = _cell(row, class_col)
        if not label or str(label).strip().lower() in ("total", "others"):
            # "Total" is a row sum, not a class. PIC's (and possibly other
            # clients') sheet ALSO carries a separate "OTHERS" convenience
            # row that duplicates Bonds + Engineering + Marine and Aviation
            # exactly (verified against PIC's real figures) — skip it too,
            # or those classes get double-counted into our own "Others"
            # bucket via _bucket().
            continue
        bucket = _bucket(label)
        totals[bucket]["gross_upr"] += float(_cell(row, gross_upr_col) or 0.0)
        totals[bucket]["net_upr"]   += float(_cell(row, net_upr_col) or 0.0)
        totals[bucket]["gross_dac"] += float(_cell(row, gross_dac_col) or 0.0)
        totals[bucket]["net_dac"]   += float(_cell(row, net_dac_col) or 0.0)

    return totals


def load_upr_dac_granular(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, Dict[str, float]]:
    """
    Same source data as load_upr_dac(), aggregated into the 6-class
    breakdown (GRANULAR_CLASSES). Unlike OCR, PIC's UPR & DAC sheet already
    stores Bonds/Engineering/Marine as their own rows — no per-client
    mapping is needed here, just label normalisation (see
    _GRANULAR_LABEL_ALIASES). Confirmed to match PIC's real published
    figures to the dollar.

    Returns:
        {"Motor": {"gross_upr":.., "net_upr":.., "gross_dac":.., "net_dac":..}, ..., "Unclassified": {...}}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "upr_dac_workbook", client.data_files["upr_dac_sheet"])

    header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if header_idx is None:
        raise ValueError("Could not locate the 'Class of Business' header row in the UPR & DAC sheet")
    gross_upr_col = class_col + 1
    net_upr_col   = class_col + 2
    gross_dac_col = class_col + 3
    net_dac_col   = class_col + 4

    totals = {c: {"gross_upr": 0.0, "net_upr": 0.0, "gross_dac": 0.0, "net_dac": 0.0} for c in GRANULAR_CLASSES + ["Unclassified"]}
    for row in rows[header_idx + 1:]:
        label = _cell(row, class_col)
        if not label or str(label).strip().lower() in ("total", "others"):
            continue
        bucket = _bucket_granular(label, client.ocr_class_mapping)
        totals[bucket]["gross_upr"] += float(_cell(row, gross_upr_col) or 0.0)
        totals[bucket]["net_upr"]   += float(_cell(row, net_upr_col) or 0.0)
        totals[bucket]["gross_dac"] += float(_cell(row, gross_dac_col) or 0.0)
        totals[bucket]["net_dac"]   += float(_cell(row, net_dac_col) or 0.0)

    return totals


# ── 4. ULAE, by class ────────────────────────────────────────────────────────

def load_ulae(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, float]:
    """
    Read the client's own computed ULAE (Alpha-ratio method) by class,
    aggregated into the 4 reserving buckets. ULAE is assumed not reinsured
    (Net ULAE = Gross ULAE, RI ULAE = 0), following PIC's own reporting
    convention — the NIC summary applies that convention for every client.

    Returns:
        {"Motor": float, "Fire": float, "Accident": float, "Others": float}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "ulae_workbook", client.data_files["ulae_sheet"])

    header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if header_idx is None:
        raise ValueError("Could not locate the 'Class of Business' header row in the ULAE sheet")
    ulae_col = _find_col(rows[header_idx], "ULAE", mode="exact")
    if ulae_col is None:
        raise ValueError("Could not locate the 'ULAE' header column in the ULAE sheet")

    totals = {c: 0.0 for c in RESERVING_CLASSES}
    for row in rows[header_idx + 1:]:
        label = _cell(row, class_col)
        if not label or str(label).strip().lower() == "total":
            continue
        bucket = _bucket(label)
        totals[bucket] += float(_cell(row, ulae_col) or 0.0)

    return totals


def load_ulae_granular(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, float]:
    """
    Same source data as load_ulae(), read at 6-class resolution
    (GRANULAR_CLASSES). NOTE: PIC's ULAE workbook only carries Bonds/
    Engineering/Marine as a single combined "Others" row (no individual
    breakdown exists in the source) — this function returns 0.0 for those
    three classes, NOT PIC's real per-class split. Callers needing a
    non-zero estimate for those three must allocate load_ulae()'s "Others"
    total themselves — see engine/ifrs17_nonlife.py's
    allocate_others_across_granular_classes(), which does this using each
    class's OCR+UPR share as a disclosed proxy (PIC's real split reflects
    actuarial judgement not visible in the source data — see that
    function's docstring).

    Returns:
        {"Motor": float, "Fire": float, "Accident": float, "Bonds": 0.0, "Engineering": 0.0, "Marine": 0.0, "Unclassified": float}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "ulae_workbook", client.data_files["ulae_sheet"])

    header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if header_idx is None:
        raise ValueError("Could not locate the 'Class of Business' header row in the ULAE sheet")
    ulae_col = _find_col(rows[header_idx], "ULAE", mode="exact")
    if ulae_col is None:
        raise ValueError("Could not locate the 'ULAE' header column in the ULAE sheet")

    totals = {c: 0.0 for c in GRANULAR_CLASSES + ["Unclassified"]}
    for row in rows[header_idx + 1:]:
        label = _cell(row, class_col)
        if not label or str(label).strip().lower() in ("total", "others"):
            continue
        bucket = _bucket_granular(label, client.ocr_class_mapping)
        totals[bucket] += float(_cell(row, ulae_col) or 0.0)

    return totals


# ── 5. Paid claims, by class (for journal "claims paid" movements) ─────────

def load_paid_claims(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, float]:
    """
    Read the "Paid Claims in <year>" column from the same ULAE workbook
    sheet used by load_ulae() — actual cash claims paid during the year,
    by class, aggregated into the 4 reserving buckets. This is a cash flow
    figure (distinct from IBNR/OCR, which are unpaid reserve estimates).

    Returns:
        {"Motor": float, "Fire": float, "Accident": float, "Others": float}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "ulae_workbook", client.data_files["ulae_sheet"])

    header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if header_idx is None:
        raise ValueError("Could not locate the 'Class of Business' header row in the ULAE sheet")
    paid_col = _find_col(rows[header_idx], "Paid Claims", mode="startswith")
    if paid_col is None:
        raise ValueError("Could not locate the 'Paid Claims' header column in the ULAE sheet")

    totals = {c: 0.0 for c in RESERVING_CLASSES}
    for row in rows[header_idx + 1:]:
        label = _cell(row, class_col)
        if not label or str(label).strip().lower() == "total":
            continue
        bucket = _bucket(label)
        totals[bucket] += float(_cell(row, paid_col) or 0.0)

    return totals


def load_paid_claims_granular(client_id: str = "pic", data_folder_override: Optional[str] = None) -> Dict[str, float]:
    """
    Same source data as load_paid_claims(), read at 6-class resolution.
    Same caveat as load_ulae_granular(): Bonds/Engineering/Marine come back
    as 0.0 here since PIC's source only carries a combined "Others" figure
    for paid claims too — allocate load_paid_claims()'s "Others" total via
    allocate_others_across_granular_classes() if a non-zero estimate is needed.

    Returns:
        {"Motor": float, "Fire": float, "Accident": float, "Bonds": 0.0, "Engineering": 0.0, "Marine": 0.0, "Unclassified": float}
    """
    client = _resolve_client(client_id, data_folder_override)
    rows = _read_sheet_rows(client, "ulae_workbook", client.data_files["ulae_sheet"])

    header_idx, class_col = _find_row_with_col(rows, "Class of Business", mode="exact")
    if header_idx is None:
        raise ValueError("Could not locate the 'Class of Business' header row in the ULAE sheet")
    paid_col = _find_col(rows[header_idx], "Paid Claims", mode="startswith")
    if paid_col is None:
        raise ValueError("Could not locate the 'Paid Claims' header column in the ULAE sheet")

    totals = {c: 0.0 for c in GRANULAR_CLASSES + ["Unclassified"]}
    for row in rows[header_idx + 1:]:
        label = _cell(row, class_col)
        if not label or str(label).strip().lower() in ("total", "others"):
            continue
        bucket = _bucket_granular(label, client.ocr_class_mapping)
        totals[bucket] += float(_cell(row, paid_col) or 0.0)

    return totals


# ── 6. RBC / GIRBC solvency data (engine/rbc/) ──────────────────────────────
#
# Reads a client's real GIRBC official return workbook (see
# ClientConfig.rbc_data_folder / rbc_data_files — a SEPARATE folder/file
# from the reserving data above, see that field's docstring for why) and
# returns the five engine.rbc.data_model dataclasses the RBC risk modules
# need. Built and verified directly against QIC's real 2025 filing
# (QIC_GIRBC_Official_Return_2025.xlsx) — every label string below is a
# real header/row label read from that workbook's GIRBC2/3/4/5/6 sheets,
# not guessed.
#
# Known approximations — read before trusting this against a different
# client's workbook, or before treating a client's Market/Credit figures as
# fully reconciled:
#   - InsuranceRiskExposures: the real GIRBC3 sheet groups classes into 9
#     named "Segments" which roll up into 5 "Categories" (Liability-like,
#     Motor-like, Property-like, Other, Credit). This loader reads the
#     Category-level Summary totals (reliable, clean labels) and assigns
#     each Category's total premium/reserve to ONE representative class in
#     engine.rbc.insurance_risk's 11-class taxonomy (Liability-like->
#     "Liability", Motor-like->"Motor", Property-like->"Fire", Other->
#     "Accident", Credit->"Miscellaneous") rather than splitting further —
#     mathematically this gives the SAME total_insurance_risk_scr as a true
#     split would (a segment with only one non-zero class just returns that
#     class's own charge unchanged), it just doesn't produce a genuinely
#     per-class breakdown in the UI.
#   - MarketRiskExposures' Equity section: the real template has 7 equity
#     categories with 7 different factors; this data model only has 4
#     ("domestic"/"foreign_developed"/"foreign_emerging"/"unlisted"). This
#     loader maps the 3 categories with a clean 1:1 conceptual match
#     (GSE-listed/developed/emerging) directly, and sums the remaining 4
#     real categories (hybrid debt, two related-party buckets, other
#     equity — which carry DIFFERENT real factors: 20%/40%/50%/50%) into
#     "unlisted". This will NOT exactly reproduce a real Equity charge
#     whenever those 4 categories are non-zero (as they are for QIC) — see
#     engine/rbc/market_risk.py's own module docstring, which already flags
#     this factor-table mismatch as needing correction; this loader
#     approximation is a direct consequence of that, not a separate issue.
#   - CreditRiskExposures "Other" bucket: the real GIRBC5 sheet has more
#     distinct line items (e.g. current tax assets at 0%, mandatory pool
#     receivables at 0.7%) than this data model's 6 fields. Receivables
#     aged 60+ days and salvage/subrogation recoverables are both folded
#     into other_receivables here (both carry the real 20% factor, so this
#     specific collapse doesn't lose accuracy) — but mandatory pool
#     receivables and current tax assets have no field to go into and are
#     NOT currently read by this loader (a real, disclosed gap — flagged
#     for Phase 4 rather than silently dropped).
#   - Counterparty/mortgage exposures (RC-rated and LTV-banded) are parsed
#     best-effort but UNTESTED against real non-zero data — QIC's own 2025
#     filing left both sections entirely blank.

from engine.rbc.data_model import (
    CreditRiskExposures, InsuranceRiskExposures, MarketRiskExposures,
    OperationalRiskExposures, QualifyingCapitalResources,
)

# GIRBC3 Category -> this module's representative class (see docstring above).
_GIRBC3_CATEGORY_TO_CLASS = {
    "Liability-like": "Liability", "Motor-like": "Motor", "Property-like": "Fire",
    "Other": "Accident", "Credit": "Miscellaneous",
}

_MORTGAGE_LTV_LABELS = {
    "LTV ≤ 40%": "<50%", "LTV ≤ 50%": "<50%", "40% < LTV ≤ 60%": "50-60%",
    "50% < LTV ≤ 60%": "50-60%", "60% < LTV ≤ 80%": "60-70%", "80% < LTV ≤ 90%": "80-90%",
    "90% < LTV ≤ 100%": ">90%", "LTV > 100%": ">90%",
}


def _value_right_of_label(rows: List[tuple], label: str, mode: str = "contains", occurrence: int = 0) -> Optional[float]:
    """
    Find the `occurrence`-th row containing a cell matching `label`, and
    return the first numeric value in that row AFTER the label's column.
    Returns None if not found or the value isn't numeric.
    """
    seen = 0
    for row in rows:
        col = _find_col(row, label, mode)
        if col is None:
            continue
        if seen < occurrence:
            seen += 1
            continue
        for v in row[col + 1:]:
            if isinstance(v, (int, float)):
                return float(v)
        return None
    return None


def _rbc_workbook_rows(client: ClientConfig, sheet_name: str, workbook_key: str = "girbc_workbook") -> List[tuple]:
    if not client.rbc_data_folder:
        raise ValueError(
            f"Client '{client.client_id}' has no rbc_data_folder configured — set the "
            f"{client.client_id.upper()}_RBC_DATA_DIR environment variable, or rbc_data_folder "
            f"in clients/{client.client_id}/client.yaml for local development."
        )
    if workbook_key not in client.rbc_data_files:
        raise ValueError(f"Client '{client.client_id}' has no '{workbook_key}' configured in rbc_data_files")
    path = os.path.join(client.rbc_data_folder, client.rbc_data_files[workbook_key])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {path} — available: {wb.sheetnames}")
        return list(wb[sheet_name].iter_rows(values_only=True))
    finally:
        wb.close()


def _parse_girbc2_capital_resources(rows: List[tuple]) -> QualifyingCapitalResources:
    a = _value_right_of_label(rows, "Tier 1 unlimited financial instruments [A]") or 0.0
    b = _value_right_of_label(rows, "Tier 1 capital components other than financial instruments [B]") or 0.0
    c = _value_right_of_label(rows, "Deductions from Tier 1 capital resources [C]") or 0.0
    e = _value_right_of_label(rows, "Tier 1 limited financial instruments [E]") or 0.0
    i = _value_right_of_label(rows, "Tier 2 financial instruments [I]") or 0.0
    j = _value_right_of_label(rows, "Tier 2 capital components other than financial instruments [J]") or 0.0
    k = _value_right_of_label(rows, "Deductions from Tier 2 capital resources [K]") or 0.0
    amortisation = _value_right_of_label(rows, "less: accumulated amortisation") or 0.0

    return QualifyingCapitalResources(
        tier1_unlimited=a + b, tier1_unlimited_deductions=c,
        tier1_limited=e, tier2=i + j - k, tier2_amortisation=amortisation,
    )


def _parse_girbc3_insurance_exposures(rows: List[tuple]) -> Tuple[InsuranceRiskExposures, float]:
    # Category names ("Liability-like" etc.) appear in THREE places in this
    # sheet: once per individual product row's Category column, once as the
    # Summary block's own row labels (what we want), and AGAIN as a
    # correlation-matrix header on the Summary block's OWN header row
    # (e.g. row 0 of the summary block has 'Liability-like' at column 16 as
    # a matrix column header, not an exposure value). A generic "find this
    # label anywhere in the row" search matches that header row first
    # (since it's scanned before the real data row) and reads a matrix
    # value instead of an exposure — anchor on the exact COLUMN "Summary"
    # itself was found in, not just any column, to avoid it.
    summary_idx, summary_col = _find_row_with_col(rows, "Summary", mode="exact")
    summary_rows = rows[summary_idx:] if summary_idx is not None else rows

    net_premium: Dict[str, float] = {}
    net_claims_reserve: Dict[str, float] = {}
    if summary_col is not None:
        for row in summary_rows:
            label = _cell(row, summary_col)
            cls = _GIRBC3_CATEGORY_TO_CLASS.get(label)
            if cls is None:
                continue
            exposure = _cell(row, summary_col + 1)
            reserve = _cell(row, summary_col + 2)
            if isinstance(exposure, (int, float)):
                net_premium[cls] = net_premium.get(cls, 0.0) + float(exposure)
            if isinstance(reserve, (int, float)):
                net_claims_reserve[cls] = net_claims_reserve.get(cls, 0.0) + float(reserve)

    revenue = _value_right_of_label(
        rows, "Total insurance revenue less the reinsurance premiums paid for the 12 months preceding the reporting date"
    ) or sum(net_premium.values())

    return InsuranceRiskExposures(net_premium=net_premium, net_claims_reserve=net_claims_reserve), revenue


def _parse_girbc4_market_exposures(rows: List[tuple]) -> MarketRiskExposures:
    ir_assets = _value_right_of_label(rows, "Present value of interest sensitive assets", occurrence=0) or 0.0
    ir_liabilities = _value_right_of_label(rows, "Present value of interest sensitive liabilities", occurrence=0) or 0.0

    fx: Dict[str, float] = {}
    for ccy in ("EURO", "GBP", "USD", "Other"):
        row_idx, col = _find_row_with_col(rows, ccy, mode="exact")
        if row_idx is None:
            continue
        row = rows[row_idx]
        long_pos = _cell(row, col + 3)
        short_pos = _cell(row, col + 4)
        net = (long_pos or 0.0) - (short_pos or 0.0)
        if net:
            fx[ccy] = float(net)

    # 7 real categories mapped 1:1 to engine.rbc.market_risk.EQUITY_FACTORS'
    # 7 keys (corrected 2026-08-03 — previously 4 of these were collapsed
    # into "unlisted", which lost the distinct hybrid_debt/related_party
    # factors; see that module's docstring).
    equity_label_map = {
        "Equity listed on the Ghana Stock Exchange": "domestic",
        "Listed equity in developed markets": "foreign_developed",
        "Listed equity in emerging markets": "foreign_emerging",
        "Hybrid debt and preference shares": "hybrid_debt",
        "Equity investments in related parties that are prudentially regulated financial institutions": "related_party_regulated",
        "Equity investments in related parties that are not prudentially regulated financial institutions": "related_party_unregulated",
        "Other equity investments": "unlisted",
    }
    listed_equities: Dict[str, float] = {}
    for label, cat in equity_label_map.items():
        v = _value_right_of_label(rows, label, mode="startswith")
        if v:
            listed_equities[cat] = listed_equities.get(cat, 0.0) + v

    real_estate_label_map = {
        "Real estate for own use": "domestic",
        "Real estate investments in related parties that are not prudentially regulated financial institutions": "foreign",
        "Other real estate investments that generate material predictable revenues": "foreign",
        "Other real estate investments": "foreign",
    }
    real_estate: Dict[str, float] = {}
    for label, cat in real_estate_label_map.items():
        v = _value_right_of_label(rows, label, mode="startswith")
        if v:
            real_estate[cat] = real_estate.get(cat, 0.0) + v

    rou_label_map = {
        "Right-of-use assets associated with leased owner-occupied properties": "owner_occupied",
        'Right-of-use assets associated with leased "other assets"': "other_assets",
        "Right-of-use assets associated with leased investment properties": "investment_property",
    }
    right_of_use: Dict[str, float] = {}
    for label, cat in rou_label_map.items():
        v = _value_right_of_label(rows, label, mode="startswith")
        if v:
            right_of_use[cat] = right_of_use.get(cat, 0.0) + v

    return MarketRiskExposures(
        listed_equities=listed_equities, real_estate=real_estate, right_of_use_assets=right_of_use,
        fx_net_open_position=fx,
        interest_rate_sensitive_assets=ir_assets, interest_rate_sensitive_liabilities=ir_liabilities,
    )


def _parse_girbc5_credit_exposures(rows: List[tuple]) -> CreditRiskExposures:
    other_map = {
        "Cash held on the insurer": "cash_and_deposits",
        "Receivables, outstanding less than 60 days": "premium_receivables",
        "Receivables, outstanding 60 days or more": "other_receivables",
        "Other recoverables (mainly salvage and subrogation)": "other_receivables",
        "Deferred tax assets arising from temporary differences": "deferred_tax_assets",
        "Loans or other forms of lending": "related_party_loans",
        "Insurance receivables from reinsurers": "reinsurance_recoverables",
        "Receivables from mandatory insurance pools": "mandatory_pool_recoverables",
    }
    totals: Dict[str, float] = {}
    for label, field_name in other_map.items():
        v = _value_right_of_label(rows, label, mode="startswith")
        if v:
            totals[field_name] = totals.get(field_name, 0.0) + v

    counterparty_exposures: List[Tuple[float, str]] = []
    for rating in ("RC1", "RC2", "RC3", "RC4", "RC5", "RC6", "RC7", "Unrated", "Default"):
        row_idx, col = _find_row_with_col(rows, rating, mode="exact")
        if row_idx is None:
            continue
        v = _cell(rows[row_idx], col + 1)
        if isinstance(v, (int, float)) and v:
            counterparty_exposures.append((float(v), rating))

    mortgage_exposures: List[Tuple[float, str]] = []
    for real_label, ltv_band in _MORTGAGE_LTV_LABELS.items():
        row_idx, col = _find_row_with_col(rows, real_label, mode="exact")
        if row_idx is None:
            continue
        v = _cell(rows[row_idx], col + 1)
        if isinstance(v, (int, float)) and v:
            mortgage_exposures.append((float(v), ltv_band))

    return CreditRiskExposures(
        counterparty_exposures=counterparty_exposures, mortgage_exposures=mortgage_exposures,
        cash_and_deposits=totals.get("cash_and_deposits", 0.0),
        premium_receivables=totals.get("premium_receivables", 0.0),
        reinsurance_recoverables=totals.get("reinsurance_recoverables", 0.0),
        mandatory_pool_recoverables=totals.get("mandatory_pool_recoverables", 0.0),
        deferred_tax_assets=totals.get("deferred_tax_assets", 0.0),
        related_party_loans=totals.get("related_party_loans", 0.0),
        other_receivables=totals.get("other_receivables", 0.0),
    )


def _parse_girbc6_operational_exposures(rows: List[tuple]) -> OperationalRiskExposures:
    current_premium = _value_right_of_label(rows, "Total insurance revenue in the most recent financial year", occurrence=0) or 0.0
    prior_premium = _value_right_of_label(rows, "Total insurance revenue in the year previous to the most recent financial year") or 0.0
    current_liabilities = _value_right_of_label(rows, "Gross current estimate at the reporting date") or 0.0

    return OperationalRiskExposures(
        current_year_net_premium=current_premium, prior_year_net_premium=prior_premium,
        current_year_net_liabilities=current_liabilities, prior_year_net_liabilities=0.0,
    )


# ── 6b. Legacy solvency margin data (engine/rbc/legacy_solvency.py) ────────
#
# Reads a client's real "FCR ... Actual" workbook (MCR-SCR + Calculation
# sheets) — a DIFFERENT workbook from the GIRBC official return above, but
# the SAME rbc_data_folder (both live under a client's FCR/<year>/ tree).
# Built and verified directly against QIC's real 2025 legacy solvency
# filing — every label below is real, and the resulting Legacy CAR
# reproduces the real 117.37% ("117%") to 5 decimal places.

_ASSET_DISCOUNT_LABEL_MAP = {
    "GOG Securities": "gog_securities",
    "Bank of Ghana Securities": "bog_securities",
    "Statutory Deposit": "statutory_deposit",
    "Cash and Term Deposits with a licensed bank": "cash_and_term_deposits",
    "Corporate Debt": "corporate_debt",
    "Quoted Equity Securities (GSE)": "listed_equities_gse",
    "Other Securities": "other_securities",
    "Equity Backed Mutual Funds": "equity_backed_mutual_funds",
    "Money Market Mutual Funds": "money_market_mutual_funds",
    "Land and Building - Investments": "property_investment",
    "Land and Building - Own use": "property_own_use",
    "Plant, Equipment & Furniture": "plant_equipment_furniture",
    "Motor Vehicles": "motor_vehicles",
    "ICT (Hardware + Software)": "ict",
    "Amount due reinsurers < 6 months": "reinsurance_recoverables_under_6mo",
    "Any other Asset except intangible and inadmissible assets": "other_assets",
}


def _parse_legacy_solvency_inputs(mcr_scr_rows: List[tuple], calculation_rows: List[tuple]):
    from engine.rbc.data_model import LegacySolvencyInputs

    core_gross = _value_right_of_label(mcr_scr_rows, "Total Core Capital", mode="exact") or 0.0
    noncore_gross = _value_right_of_label(mcr_scr_rows, "Total Non-Core Capital", mode="exact") or 0.0
    total_capital_base = core_gross + noncore_gross

    core_row_idx, core_col = _find_row_with_col(mcr_scr_rows, "Total Core Capital", mode="exact")
    core_admissible = _cell(mcr_scr_rows[core_row_idx], core_col + 3) if core_row_idx is not None else 0.0
    noncore_row_idx, noncore_col = _find_row_with_col(mcr_scr_rows, "Total Non-Core Capital", mode="exact")
    noncore_admissible = _cell(mcr_scr_rows[noncore_row_idx], noncore_col + 3) if noncore_row_idx is not None else 0.0
    capital_deductions = (core_gross - (core_admissible or 0.0)) + (noncore_gross - (noncore_admissible or 0.0))

    asset_balances: Dict[str, float] = {}
    for real_label, key in _ASSET_DISCOUNT_LABEL_MAP.items():
        row_idx, col = _find_row_with_col(calculation_rows, real_label, mode="exact")
        if row_idx is None:
            continue
        v = _cell(calculation_rows[row_idx], col + 2)
        asset_balances[key] = float(v) if isinstance(v, (int, float)) else 0.0

    net_written_premium = _value_right_of_label(calculation_rows, "Net Written Premium (current year)", mode="exact") or 0.0
    management_expenses = _value_right_of_label(
        calculation_rows, "Non-attributable Expenses + Acquisition Cashflows", mode="exact"
    ) or 0.0

    return LegacySolvencyInputs(
        total_capital_base=total_capital_base, capital_deductions=capital_deductions,
        asset_balances=asset_balances, net_written_premium=net_written_premium,
        management_expenses=management_expenses,
    )


def load_rbc_solvency_data(client_id: str = "pic") -> dict:
    """
    Parse a client's real GIRBC official return workbook into the five
    engine.rbc.data_model inputs, PLUS (if the client has a legacy_workbook
    configured in rbc_data_files) the legacy solvency-margin inputs from a
    separate "FCR ... Actual" workbook. See this section's module-level
    comment (above) for known approximations vs a fully granular
    reconstruction.

    Returns:
        {
            "capital_resources":    QualifyingCapitalResources,
            "insurance_risk":       InsuranceRiskExposures,
            "net_non_life_insurance_revenue": float,
            "market_risk":            MarketRiskExposures,
            "credit_risk":               CreditRiskExposures,
            "operational_risk":             OperationalRiskExposures,
            "legacy_inputs":                   LegacySolvencyInputs | None,  # None if legacy_workbook isn't configured for this client
        }
    """
    client = load_client(client_id)

    girbc2 = _rbc_workbook_rows(client, "GIRBC2-Qual. Capital Resources")
    girbc3 = _rbc_workbook_rows(client, "GIRBC3-Insurance Risks")
    girbc4 = _rbc_workbook_rows(client, "GIRBC4-Market Risks")
    girbc5 = _rbc_workbook_rows(client, "GIRBC5-Credit Risks")
    girbc6 = _rbc_workbook_rows(client, "GIRBC6-Operational Risk")

    insurance_risk, revenue = _parse_girbc3_insurance_exposures(girbc3)

    legacy_inputs = None
    if "legacy_workbook" in client.rbc_data_files:
        mcr_scr_rows = _rbc_workbook_rows(client, "MCR-SCR", workbook_key="legacy_workbook")
        calculation_rows = _rbc_workbook_rows(client, "Calculation", workbook_key="legacy_workbook")
        legacy_inputs = _parse_legacy_solvency_inputs(mcr_scr_rows, calculation_rows)

    return {
        "capital_resources":               _parse_girbc2_capital_resources(girbc2),
        "insurance_risk":                     insurance_risk,
        "net_non_life_insurance_revenue":        revenue,
        "market_risk":                              _parse_girbc4_market_exposures(girbc4),
        "credit_risk":                                  _parse_girbc5_credit_exposures(girbc5),
        "operational_risk":                                _parse_girbc6_operational_exposures(girbc6),
        "legacy_inputs":                                      legacy_inputs,
    }


# ── Quick test ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    """
    Run this file directly to sanity-check the parser against the live files:
        cd amvs
        python -m engine.data_loader [client_id]
    """
    import sys
    cid = sys.argv[1] if len(sys.argv) > 1 else "pic"

    for cls in RESERVING_CLASSES:
        tri = load_triangle(cls, client_id=cid)
        print(f"{cls}: {len(tri['gross_triangle'])} origin years, "
              f"latest gross={list(tri['gross_triangle'].items())[-1]}")

    print("\nOCR:", load_ocr(client_id=cid))
    print("\nUPR/DAC:", load_upr_dac(client_id=cid))
    print("\nULAE:", load_ulae(client_id=cid))
