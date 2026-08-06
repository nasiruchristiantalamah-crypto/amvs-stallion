"""
================================================================================
CASH FLOW BY COVERED LIFE — per-dependant regulatory audit trail
================================================================================
What this file does:
    Validates the per-life cash flow breakdown added for the Excel export's
    "Cash Flow by Covered Life" sheet — engine/cashflows.py's
    CashFlowRow.benefit_lines_by_life, and engine/custom_pricing.py's
    _annual_cashflow_by_life_rollup built on top of it.

    The core invariant throughout: this is the SAME computed cash flow
    data, just attributed to who it's actually for — summing every life's
    contribution for a given rider (month) or (rider, year) must always
    reproduce the existing combined benefit_lines / claims_by_rider figure
    exactly. If that ever stops being true, the per-life breakdown has
    drifted from the number actually used to solve the premium — the one
    failure mode that would make the Excel sheet misleading rather than
    just incomplete.
================================================================================
"""

import pytest

from api.main import CustomProductRequest, _build_product_from_request, _resolve_custom_assumptions
from engine.assumptions import ProductAssumptions, LapseSchedule
from engine.product import Product, Rider, Dependant
from engine.decrement import run_decrement_projection
from engine.cashflows import calculate_cash_flows, _dependant_life_labels
from engine.custom_pricing import run_custom_pricing


def _product_with_dependants():
    return Product(name="Test", riders=[
        Rider(rider_type="death", name="Main Benefit", benefit_main=50000, benefit_dependant=50000, incidence_basis="mortality"),
        Rider(rider_type="tpd", name="TPD", benefit_main=25000, benefit_dependant=25000,
              incidence_basis="mortality_multiple", mortality_multiplier=0.2),
    ], dependants=[
        Dependant(relationship="spouse", age=38),
        Dependant(relationship="child", age=10, benefit_multiplier=0.5),
    ])


# ── Life labels ───────────────────────────────────────────────────────────

def test_life_labels_use_relationship_names():
    product = _product_with_dependants()
    assert _dependant_life_labels(product) == ["Spouse", "Child"]


def test_duplicate_relationships_get_disambiguated():
    product = Product(name="Test", dependants=[
        Dependant(relationship="child", age=8),
        Dependant(relationship="child", age=12),
    ])
    assert _dependant_life_labels(product) == ["Child 1", "Child 2"]


def test_no_dependants_gives_no_labels():
    product = Product(name="Test")
    assert _dependant_life_labels(product) == []


# ── engine/cashflows.py: benefit_lines_by_life ───────────────────────────────

def test_benefit_lines_by_life_covers_main_life_and_every_dependant():
    product = _product_with_dependants()
    assumptions = ProductAssumptions(entry_age_main=35, lapse_schedule=LapseSchedule(rates={1: 0.0}))
    dec_rows = run_decrement_projection(assumptions, product)
    cf_rows = calculate_cash_flows(dec_rows, assumptions, product, monthly_premium=50.0)

    month_one = cf_rows[0]
    assert set(month_one.benefit_lines_by_life.keys()) == {"Main Life", "Spouse", "Child"}


def test_benefit_lines_by_life_sums_to_the_combined_total_every_month():
    product = _product_with_dependants()
    assumptions = ProductAssumptions(entry_age_main=35, lapse_schedule=LapseSchedule(rates={1: 0.30, 13: 0.02}))
    dec_rows = run_decrement_projection(assumptions, product)
    cf_rows = calculate_cash_flows(dec_rows, assumptions, product, monthly_premium=50.0)

    # Check every month, not just month 1 — the invariant must hold as
    # lives decrement and riders hit waiting periods/expiry, not just at
    # the clean starting point.
    for cf in cf_rows:
        for rider_name, combined_amount in cf.benefit_lines.items():
            life_sum = sum(cf.benefit_lines_by_life[life].get(rider_name, 0.0) for life in cf.benefit_lines_by_life)
            assert life_sum == pytest.approx(combined_amount, abs=1e-9), \
                f"month {cf.month}, rider {rider_name}: combined={combined_amount}, life_sum={life_sum}"


def test_dependant_with_zero_override_gets_no_benefit_lines_entry():
    # A dependant with benefit_overrides = {"TPD": 0} should contribute
    # nothing to TPD anywhere, including their own per-life breakdown.
    product = Product(name="Test", riders=[
        Rider(rider_type="tpd", name="TPD", benefit_main=25000, benefit_dependant=25000,
              incidence_basis="mortality_multiple", mortality_multiplier=0.2),
    ], dependants=[
        Dependant(relationship="spouse", age=38, benefit_overrides={"TPD": 0.0}),
    ])
    assumptions = ProductAssumptions(entry_age_main=35, lapse_schedule=LapseSchedule(rates={1: 0.0}))
    dec_rows = run_decrement_projection(assumptions, product)
    cf_rows = calculate_cash_flows(dec_rows, assumptions, product, monthly_premium=50.0)
    assert "TPD" not in cf_rows[0].benefit_lines_by_life["Spouse"]


# ── engine/custom_pricing.py: annual rollup ──────────────────────────────────

def test_annual_rollup_reconciles_to_the_combined_annual_rollup():
    req = CustomProductRequest(
        product_name="Test", product_type="micro_life", sum_assured=50000, entry_age=35,
        riders=[{"name": "TPD", "benefit_type": "tpd", "benefit_amount": 25000,
                 "waiting_period_months": 0, "mortality_multiplier": 0.2}],
        dependants=[{"relationship": "spouse", "age": 38}, {"relationship": "child", "age": 10}],
    )
    product = _build_product_from_request(req)
    assumptions = _resolve_custom_assumptions(req)
    result = run_custom_pricing(product, assumptions)

    by_life = result["annual_cashflow_by_life"]
    combined = result["annual_cashflow"]
    assert set(by_life.keys()) == {"Main Life", "Spouse", "Child"}

    for year_combined in combined:
        y = year_combined["policy_year"]
        life_sum = sum(
            next(r["total_claims"] for r in by_life[life] if r["policy_year"] == y)
            for life in by_life
        )
        assert life_sum == pytest.approx(year_combined["total_claims"], abs=0.1), f"policy year {y}"


def test_annual_rollup_tracks_decrement_not_just_cashflow():
    req = CustomProductRequest(product_name="Test", product_type="whole_life", sum_assured=50000, entry_age=35,
                                dependants=[{"relationship": "spouse", "age": 38}])
    product = _build_product_from_request(req)
    assumptions = _resolve_custom_assumptions(req)
    result = run_custom_pricing(product, assumptions)

    spouse_rows = result["annual_cashflow_by_life"]["Spouse"]
    # Opening lives must start at (or very near) 1.0 and decline over time
    # as the spouse's own decrement plays out.
    assert spouse_rows[0]["opening_lives"] == pytest.approx(1.0, abs=1e-6)
    assert spouse_rows[-1]["opening_lives"] < spouse_rows[0]["opening_lives"]
    assert spouse_rows[0]["expected_deaths"] > 0
    assert spouse_rows[0]["expected_lapses"] > 0


def test_excel_export_includes_cashflow_by_life_sheet():
    from outputs.custom_pricing_excel_exporter import export_custom_pricing_to_excel
    import openpyxl
    import os

    req = CustomProductRequest(
        product_name="Excel Export Test", product_type="micro_life", sum_assured=50000, entry_age=35,
        dependants=[{"relationship": "spouse", "age": 38}],
    )
    product = _build_product_from_request(req)
    assumptions = _resolve_custom_assumptions(req)
    result = run_custom_pricing(product, assumptions)

    path = export_custom_pricing_to_excel(result, req.model_dump())
    try:
        wb = openpyxl.load_workbook(path)
        assert "Cash Flow by Covered Life" in wb.sheetnames
        ws = wb["Cash Flow by Covered Life"]
        # Section headers for both lives must appear somewhere in the sheet.
        all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
        assert "MAIN LIFE" in all_values
        assert "SPOUSE" in all_values
    finally:
        os.remove(path)
