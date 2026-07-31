"""
================================================================================
EXCEL EXPORTER VALIDATION
================================================================================
Validates outputs/excel_exporter.py end-to-end with real PIC-derived data:
produces an actual workbook, confirms all 7 required sheets exist with
Stallion navy formatting, and cross-checks a few figures in the exported
cells against the underlying statements/journal objects (not just that a
file was written, but that the numbers in it are the real numbers).

Run with:
    cd amvs
    pytest tests/test_excel_exporter.py -s
================================================================================
"""

import os

import openpyxl
import pytest

from engine.ifrs17_nonlife import generate_nonlife_paa_statements
from engine.data_loader import load_paid_claims
from engine.journals import generate_nonlife_journal
from outputs.excel_exporter import export_nonlife_statements_to_excel, STALLION_NAVY, GHS_FORMAT

REQUIRED_SHEETS = [
    "Summary", "IBNR by class", "PAA liabilities", "Balance sheet",
    "Income statement", "Journal entries", "Assumptions",
]


@pytest.fixture(scope="module")
def exported(tmp_path_factory):
    statements = generate_nonlife_paa_statements(verbose=False)
    paid = load_paid_claims()
    entries = generate_nonlife_journal(statements, paid, period="FY2025")

    out_dir = tmp_path_factory.mktemp("excel_export")
    out_path = os.path.join(str(out_dir), "test_nonlife_statements.xlsx")
    written_path = export_nonlife_statements_to_excel(
        statements, entries, meta={"company_name": "Provident Insurance Limited"}, output_path=out_path,
    )
    return {"path": written_path, "statements": statements, "entries": entries}


def test_file_is_written(exported):
    assert os.path.isfile(exported["path"])
    assert os.path.getsize(exported["path"]) > 5000


def test_all_required_sheets_present(exported):
    wb = openpyxl.load_workbook(exported["path"])
    for sheet_name in REQUIRED_SHEETS:
        assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"


def test_header_uses_stallion_navy(exported):
    wb = openpyxl.load_workbook(exported["path"])
    ws = wb["Summary"]
    header_cell = ws["A5"]  # "Basis" header of the headline liabilities table
    assert header_cell.fill.start_color.rgb.endswith(STALLION_NAVY)


def test_currency_cells_use_ghs_format(exported):
    wb = openpyxl.load_workbook(exported["path"])
    ws = wb["Summary"]
    assert ws["B6"].number_format == GHS_FORMAT   # Gross LRC figure


def test_summary_totals_match_statements(exported):
    wb = openpyxl.load_workbook(exported["path"])
    ws = wb["Summary"]
    totals = exported["statements"]["totals"]
    assert ws["B6"].value == totals["gross"].lrc
    assert ws["C6"].value == totals["gross"].lic
    assert ws["D6"].value == totals["gross"].total_liability


def test_journal_sheet_totals_match_and_balance(exported):
    wb = openpyxl.load_workbook(exported["path"])
    ws = wb["Journal entries"]
    last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
    excel_total_debit  = last_row[5]
    excel_total_credit = last_row[6]
    assert excel_total_debit == excel_total_credit

    expected_debit = round(sum(e.debit for e in exported["entries"]), 2)
    assert excel_total_debit == expected_debit
