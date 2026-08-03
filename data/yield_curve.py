"""
================================================================================
NIC RISK-FREE RATE (RFR) YIELD CURVE
================================================================================
What this file does:
    Loads the National Insurance Commission's published risk-free rate term
    structure, used to discount IBNR/OCR claims reserves to present value
    for the non-life Liability for Incurred Claims (LIC) — IFRS 17 Para 36
    / B72-B85, the same "bottom-up" discount methodology referenced on the
    life side of this engine (engine/assumptions.py's valuation_rate_pa).

SOURCE:
    data/nic_curves/20251231_NIC_RFR_16012025.xlsx — committed to this repo
    (see .gitignore's explicit exception for this one file: everything else
    matching *.xlsx is excluded as generated report output, but this is a
    genuine input, and a PUBLIC one — the NIC's own published regulatory
    curve, not private client data — so unlike clients/<id>/'s real Excel
    workbooks it's safe to ship with the app itself rather than depend on
    a per-machine path or an env var). This used to be an absolute path
    into one developer's local OneDrive (a real production bug: that path
    can't exist on Railway or any other machine, and reserving would fail
    with a bare FileNotFoundError the moment discounting was attempted —
    see git history around 2026-08-03 for the incident this fixed).
    Sheet "CurrentYearResult" — Ghana Cedis spot rate column, years 1-80,
    "Risk-free curves as of 31 December 2025", produced by the NIC (GHS
    fitted from GFIM observations; USD/GBP/EUR sourced separately and also
    available via this loader).

    This is the genuine NIC-published curve, not a client-specific copy —
    confirmed by cross-checking against the identical curve embedded in
    GLICO Insurance's own 2025 GMM valuation assumptions file
    (20251231_NIC_RFR_YieldCurve.xlsx), which carries the same figures.

    A newer curve can be dropped in without a code change by setting the
    NIC_RFR_CURVE_PATH environment variable (falls back to the bundled
    file above when unset).

Structure:
    {duration_year: annual spot rate}, e.g. {1: 0.1356, 2: 0.1498, ...}
================================================================================
"""

import os
from typing import Dict

import openpyxl

_BUNDLED_YIELD_CURVE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "nic_curves", "20251231_NIC_RFR_16012025.xlsx"
)
DEFAULT_YIELD_CURVE_PATH = os.environ.get("NIC_RFR_CURVE_PATH") or _BUNDLED_YIELD_CURVE_PATH

# 0-indexed column offset of each currency's Spot Rate column, in the
# "CurrentYearResult" sheet's row tuples (Year, GHS Spot, GHS Fwd, blank,
# USD Spot, USD Fwd, blank, GBP Spot, GBP Fwd, blank, EUR Spot, EUR Fwd).
_CURRENCY_SPOT_COLUMNS = {"GHS": 1, "USD": 4, "GBP": 7, "EUR": 10}


def load_yield_curve(path: str = DEFAULT_YIELD_CURVE_PATH, currency: str = "GHS") -> Dict[int, float]:
    """
    Read the NIC RFR yield curve and return {year: spot_rate}.

    Parameters:
        path     : workbook path (defaults to the NIC-published curve)
        currency : "GHS" (default), "USD", "GBP", or "EUR"

    Returns:
        {duration_year: annual spot rate}
    """
    if currency not in _CURRENCY_SPOT_COLUMNS:
        raise ValueError(f"Unknown currency '{currency}' — choose from {sorted(_CURRENCY_SPOT_COLUMNS)}")

    if not os.path.isfile(path):
        raise ValueError(
            f"NIC RFR yield curve not found at '{path}'. This should be the bundled "
            f"data/nic_curves/20251231_NIC_RFR_16012025.xlsx unless NIC_RFR_CURVE_PATH is "
            f"set to something else — check that environment variable if this is unexpected."
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["CurrentYearResult"]

        header_row_num = None
        for row in ws.iter_rows(values_only=False):
            if row and row[0].value == "Year":
                header_row_num = row[0].row
                break
        if header_row_num is None:
            raise ValueError(f"Could not find the 'Year' header row in {path}, sheet 'CurrentYearResult'")

        spot_col = _CURRENCY_SPOT_COLUMNS[currency]
        curve: Dict[int, float] = {}
        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            year = row[0]
            if year is None:
                break
            curve[int(year)] = float(row[spot_col])
        return curve
    finally:
        wb.close()


def discount_factor(duration_years: float, curve: Dict[int, float]) -> float:
    """
    Discount factor for a cash flow `duration_years` from now, using the
    curve's spot rate at the nearest whole year to that duration (a
    duration beyond the curve's longest published year uses that year's
    rate — the curve here runs to 80 years, so this only matters for
    unusually long-tailed liabilities).

    DF = 1 / (1 + spot_rate) ** duration_years
    """
    if not curve:
        raise ValueError("Empty yield curve — cannot compute a discount factor")
    years = sorted(curve.keys())
    nearest = min(years, key=lambda y: abs(y - duration_years))
    spot = curve[nearest]
    return 1.0 / (1.0 + spot) ** duration_years
